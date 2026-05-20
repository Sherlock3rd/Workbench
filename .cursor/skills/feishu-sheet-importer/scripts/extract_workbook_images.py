from __future__ import annotations

import argparse
import json
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Extract embedded images from an xlsx workbook.")
    parser.add_argument("--source-workbook", required=True, help="Local xlsx file.")
    parser.add_argument(
        "--output-dir",
        help="Directory to store extracted images. Defaults to a sibling folder next to the workbook.",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    workbook_path = Path(args.source_workbook).expanduser().resolve()
    wb = load_workbook(workbook_path)
    ws = wb.active

    output_dir = Path(args.output_dir).expanduser().resolve() if args.output_dir else workbook_path.with_suffix("")
    output_dir.mkdir(parents=True, exist_ok=True)

    manifest: list[dict[str, object]] = []
    for idx, image in enumerate(getattr(ws, "_images", []), start=1):
        anchor = image.anchor._from
        cell_ref = f"{get_column_letter(anchor.col + 1)}{anchor.row + 1}"
        original_name = Path(getattr(image, "path", "") or f"image{idx}.png").name
        image_name = f"{idx:02d}_{cell_ref}_{original_name}"
        target_path = output_dir / image_name
        target_path.write_bytes(image._data())
        manifest.append(
            {
                "index": idx,
                "sheet_title": ws.title,
                "cell": cell_ref,
                "row": anchor.row + 1,
                "column": anchor.col + 1,
                "filename": image_name,
                "path": str(target_path),
            }
        )

    print(json.dumps({"output_dir": str(output_dir), "images": manifest}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
