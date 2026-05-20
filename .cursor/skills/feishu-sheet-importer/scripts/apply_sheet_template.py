from __future__ import annotations

import argparse
import base64
from io import BytesIO
import json
from collections import defaultdict
from pathlib import Path
import sys
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.styles import Border, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from PIL import Image as PILImage

from common import (
    LarkCliError,
    chunk_consecutive,
    ensure_lark_success,
    excel_width_to_pixels,
    extract_spreadsheet_token,
    hex_color_from_openpyxl,
    pick_sheet_ref,
    run_lark_json,
)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Replay key workbook styles onto a Feishu spreadsheet.")
    parser.add_argument("--source-workbook", required=True, help="Local xlsx file used as the style source.")
    parser.add_argument("--spreadsheet", required=True, help="Feishu spreadsheet token or URL.")
    parser.add_argument("--identity", default="user", choices=["user", "bot"], help="lark-cli identity.")
    parser.add_argument("--sheet-title", help="Rename the first imported sheet to this title.")
    parser.add_argument("--spreadsheet-title", help="Rename the spreadsheet file title.")
    parser.add_argument(
        "--copy-merges",
        action="store_true",
        help="Replay merged ranges from the local workbook. Use when import does not preserve merged cells.",
    )
    parser.add_argument(
        "--insert-images",
        action="store_true",
        help="Replay embedded worksheet images into the target Feishu sheet.",
    )
    parser.add_argument(
        "--skip-widths",
        action="store_true",
        help="Skip column width replay.",
    )
    parser.add_argument(
        "--skip-row-heights",
        action="store_true",
        help="Skip row height replay.",
    )
    parser.add_argument(
        "--skip-styles",
        action="store_true",
        help="Skip cell style replay.",
    )
    parser.add_argument(
        "--max-style-ranges",
        type=int,
        default=500,
        help="Maximum number of grouped style ranges to send in one update pass.",
    )
    return parser.parse_args()


def maybe_patch_spreadsheet_title(spreadsheet_token: str, title: str | None, identity: str) -> None:
    if not title:
        return
    ensure_lark_success(
        run_lark_json(
            [
                "sheets",
                "spreadsheets",
                "patch",
                "--as",
                identity,
                "--params",
                json.dumps({"spreadsheet_token": spreadsheet_token}, ensure_ascii=False),
                "--data",
                json.dumps({"title": title}, ensure_ascii=False),
            ]
        )
    )


def maybe_patch_sheet_title(spreadsheet_token: str, sheet_id: str, title: str | None, identity: str) -> None:
    if not title:
        return
    body = {
        "requests": [
            {
                "updateSheet": {
                    "properties": {
                        "sheetId": sheet_id,
                        "title": title,
                    }
                }
            }
        ]
    }
    try:
        ensure_lark_success(
            run_lark_json(
                [
                    "api",
                    "POST",
                    f"/open-apis/sheets/v2/spreadsheets/{spreadsheet_token}/sheets_batch_update",
                    "--as",
                    identity,
                    "--data",
                    json.dumps(body, ensure_ascii=False),
                ]
            )
        )
    except LarkCliError as exc:
        if "sheetTitle already exist" not in (exc.stdout or "") and "sheetTitle already exist" not in (exc.stderr or ""):
            raise


def build_width_groups(ws: Worksheet) -> list[tuple[int, int, int]]:
    width_pairs: list[tuple[int, int]] = []
    for column_index in range(1, ws.max_column + 1):
        letter = get_column_letter(column_index)
        dim = ws.column_dimensions.get(letter)
        if not dim:
            continue
        width = excel_width_to_pixels(dim.width)
        if width is None:
            continue
        width_pairs.append((column_index, width))
    return chunk_consecutive(width_pairs)


def build_row_height_groups(ws: Worksheet) -> list[tuple[int, int, int]]:
    height_pairs: list[tuple[int, int]] = []
    for row_index in range(1, ws.max_row + 1):
        dim = ws.row_dimensions.get(row_index)
        if not dim:
            continue
        if dim.height is None:
            continue
        pixels = max(1, int(round(float(dim.height) * 4 / 3)))
        height_pairs.append((row_index, pixels))
    return chunk_consecutive(height_pairs)


def apply_column_widths(spreadsheet_token: str, sheet_id: str, ws: Worksheet, identity: str) -> int:
    applied = 0
    for start, end, width in build_width_groups(ws):
        body = {
            "dimension": {
                "sheetId": sheet_id,
                "majorDimension": "COLUMNS",
                "startIndex": start,
                "endIndex": end,
            },
            "dimensionProperties": {
                "fixedSize": width,
                "visible": width > 0,
            },
        }
        ensure_lark_success(
            run_lark_json(
                [
                    "api",
                    "PUT",
                    f"/open-apis/sheets/v2/spreadsheets/{spreadsheet_token}/dimension_range",
                    "--as",
                    identity,
                    "--data",
                    json.dumps(body, ensure_ascii=False),
                ]
            )
        )
        applied += 1
    return applied


def apply_row_heights(spreadsheet_token: str, sheet_id: str, ws: Worksheet, identity: str) -> tuple[int, list[str]]:
    applied = 0
    warnings: list[str] = []
    for start, end, height in build_row_height_groups(ws):
        body = {
            "dimension": {
                "sheetId": sheet_id,
                "majorDimension": "ROWS",
                "startIndex": start,
                "endIndex": end,
            },
            "dimensionProperties": {
                "fixedSize": height,
                "visible": height > 0,
            },
        }
        try:
            ensure_lark_success(
                run_lark_json(
                    [
                        "api",
                        "PUT",
                        f"/open-apis/sheets/v2/spreadsheets/{spreadsheet_token}/dimension_range",
                        "--as",
                        identity,
                        "--data",
                        json.dumps(body, ensure_ascii=False),
                    ]
                )
            )
            applied += 1
        except LarkCliError as exc:
            warnings.append(f"Skipped row heights {start}-{end}: {exc}")
    return applied, warnings


def build_style_signature(cell: Cell) -> tuple[tuple[str, Any], ...] | None:
    fill: PatternFill = cell.fill
    font: Font = cell.font
    border: Border = cell.border
    alignment = cell.alignment

    back_color = None
    if fill and getattr(fill, "fill_type", None) == "solid":
        back_color = hex_color_from_openpyxl(fill.fgColor)

    fore_color = hex_color_from_openpyxl(font.color)
    bold = bool(font.bold)
    italic = bool(font.italic)

    border_enabled = any(
        getattr(side, "style", None)
        for side in (border.left, border.right, border.top, border.bottom)
    )
    border_color = (
        hex_color_from_openpyxl(border.left.color)
        or hex_color_from_openpyxl(border.right.color)
        or hex_color_from_openpyxl(border.top.color)
        or hex_color_from_openpyxl(border.bottom.color)
    )

    h_align_map = {"left": 0, "center": 1, "centerContinuous": 1, "right": 2}
    v_align_map = {"top": 0, "center": 1, "bottom": 2}
    h_align = h_align_map.get(alignment.horizontal or "", None)
    v_align = v_align_map.get(alignment.vertical or "", None)

    signature = {
        "backColor": back_color,
        "foreColor": fore_color,
        "fontBold": bold,
        "fontItalic": italic,
        "hAlign": h_align,
        "vAlign": v_align,
        "borderType": "FULL_BORDER" if border_enabled else None,
        "borderColor": border_color if border_enabled else None,
    }
    if not any(value not in (None, False) for value in signature.values()):
        return None
    return tuple(sorted(signature.items()))


def style_signature_to_payload(signature: tuple[tuple[str, Any], ...]) -> dict[str, Any]:
    signature_dict = dict(signature)
    style: dict[str, Any] = {"clean": False}

    font_payload: dict[str, Any] = {"clean": False}
    if signature_dict.get("fontBold"):
        font_payload["bold"] = True
    if signature_dict.get("fontItalic"):
        font_payload["italic"] = True
    if len(font_payload) > 1:
        style["font"] = font_payload

    if signature_dict.get("backColor"):
        style["backColor"] = signature_dict["backColor"]
    if signature_dict.get("foreColor"):
        style["foreColor"] = signature_dict["foreColor"]
    if signature_dict.get("hAlign") is not None:
        style["hAlign"] = signature_dict["hAlign"]
    if signature_dict.get("vAlign") is not None:
        style["vAlign"] = signature_dict["vAlign"]
    if signature_dict.get("borderType"):
        style["borderType"] = signature_dict["borderType"]
    if signature_dict.get("borderColor"):
        style["borderColor"] = signature_dict["borderColor"]

    return style


def collect_style_ranges(ws: Worksheet, sheet_id: str) -> dict[tuple[tuple[str, Any], ...], list[str]]:
    grouped: dict[tuple[tuple[str, Any], ...], list[str]] = defaultdict(list)

    for row_index in range(1, ws.max_row + 1):
        active_signature: tuple[tuple[str, Any], ...] | None = None
        segment_start: int | None = None

        def flush(segment_end: int) -> None:
            nonlocal active_signature, segment_start
            if active_signature is None or segment_start is None:
                return
            start_letter = get_column_letter(segment_start)
            end_letter = get_column_letter(segment_end)
            grouped[active_signature].append(f"{sheet_id}!{start_letter}{row_index}:{end_letter}{row_index}")
            active_signature = None
            segment_start = None

        for column_index in range(1, ws.max_column + 1):
            signature = build_style_signature(ws.cell(row=row_index, column=column_index))
            if signature == active_signature:
                continue
            if active_signature is not None and segment_start is not None:
                flush(column_index - 1)
            if signature is not None:
                active_signature = signature
                segment_start = column_index
        if active_signature is not None and segment_start is not None:
            flush(ws.max_column)

    return grouped


def apply_styles(
    spreadsheet_token: str,
    ws: Worksheet,
    sheet_id: str,
    identity: str,
    max_style_ranges: int,
) -> int:
    grouped_ranges = collect_style_ranges(ws, sheet_id)
    if not grouped_ranges:
        return 0

    batches: list[dict[str, Any]] = []
    for signature, ranges in grouped_ranges.items():
        style = style_signature_to_payload(signature)
        if len(style) == 1 and style.get("clean") is False:
            continue
        for start in range(0, len(ranges), max_style_ranges):
            batches.append({"ranges": ranges[start : start + max_style_ranges], "style": style})

    if not batches:
        return 0

    for index, batch in enumerate(batches, start=1):
        try:
            ensure_lark_success(
                run_lark_json(
                    [
                        "api",
                        "PUT",
                        f"/open-apis/sheets/v2/spreadsheets/{spreadsheet_token}/styles_batch_update",
                        "--as",
                        identity,
                        "--data",
                        json.dumps({"data": [batch]}, ensure_ascii=False),
                    ]
                )
            )
        except LarkCliError as exc:
            raise LarkCliError(
                f"Style batch {index} failed.",
                stdout=exc.stdout,
                stderr=exc.stderr,
                exit_code=exc.exit_code,
            ) from exc
    return len(batches)


def apply_merges(spreadsheet_token: str, ws: Worksheet, sheet_id: str, identity: str) -> tuple[int, list[str]]:
    applied = 0
    warnings: list[str] = []
    for merged_range in ws.merged_cells.ranges:
        payload = {
            "range": f"{sheet_id}!{merged_range.coord}",
            "mergeType": "MERGE_ALL",
        }
        try:
            ensure_lark_success(
                run_lark_json(
                    [
                        "api",
                        "POST",
                        f"/open-apis/sheets/v2/spreadsheets/{spreadsheet_token}/merge_cells",
                        "--as",
                        identity,
                        "--data",
                        json.dumps(payload, ensure_ascii=False),
                    ]
                )
            )
            applied += 1
        except LarkCliError as exc:
            warnings.append(f"Skipped merge {merged_range.coord}: {exc}")
    return applied, warnings


def apply_images(spreadsheet_token: str, ws: Worksheet, sheet_id: str, identity: str) -> tuple[int, list[str]]:
    applied = 0
    warnings: list[str] = []
    for idx, image in enumerate(getattr(ws, "_images", []), start=1):
        try:
            anchor_from = image.anchor._from
            cell_ref = f"{get_column_letter(anchor_from.col + 1)}{anchor_from.row + 1}"
            image_name = Path(getattr(image, "path", "") or f"image{idx}.png").name
            image_bytes = image._data()
            with PILImage.open(BytesIO(image_bytes)) as pil_image:
                pil_image = pil_image.convert("RGB")
                pil_image.thumbnail((256, 256))
                buffer = BytesIO()
                pil_image.save(buffer, format="JPEG", quality=28, optimize=True)
                encoded_image = base64.b64encode(buffer.getvalue()).decode("ascii")

            suffix = Path(getattr(image, "path", "") or f"image{idx}.jpg").suffix or ".jpg"
            image_name = f"image_{idx:02d}{suffix.lower()}"
            body = {
                "range": f"{sheet_id}!{cell_ref}:{cell_ref}",
                "image": encoded_image,
                "name": image_name,
            }
            ensure_lark_success(
                run_lark_json(
                    [
                        "api",
                        "POST",
                        f"/open-apis/sheets/v2/spreadsheets/{spreadsheet_token}/values_image",
                        "--as",
                        identity,
                        "--data",
                        json.dumps(body, ensure_ascii=False),
                    ]
                )
            )
            applied += 1
        except Exception as exc:  # noqa: BLE001
            warnings.append(f"Skipped image {idx}: {exc}")
    return applied, warnings


def main() -> None:
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    except Exception:  # noqa: BLE001
        pass
    args = parse_args()
    workbook_path = Path(args.source_workbook).expanduser().resolve()
    wb = load_workbook(workbook_path)
    ws = wb.active

    spreadsheet_token = extract_spreadsheet_token(args.spreadsheet)
    spreadsheet_info = ensure_lark_success(
        run_lark_json(
            [
                "sheets",
                "+info",
                "--as",
                args.identity,
                "--spreadsheet-token",
                spreadsheet_token,
            ]
        )
    )
    sheet_ref = pick_sheet_ref(spreadsheet_info, preferred_title=ws.title)

    print("Applying spreadsheet title...", flush=True)
    maybe_patch_spreadsheet_title(spreadsheet_token, args.spreadsheet_title, args.identity)
    print("Applying sheet title...", flush=True)
    maybe_patch_sheet_title(spreadsheet_token, sheet_ref.sheet_id, args.sheet_title or ws.title, args.identity)
    width_updates = 0
    if not args.skip_widths:
        print("Applying column widths...", flush=True)
        width_updates = apply_column_widths(spreadsheet_token, sheet_ref.sheet_id, ws, args.identity)

    row_height_updates = 0
    row_height_warnings: list[str] = []
    if not args.skip_row_heights:
        print("Applying row heights...", flush=True)
        row_height_updates, row_height_warnings = apply_row_heights(spreadsheet_token, sheet_ref.sheet_id, ws, args.identity)

    style_updates = 0
    if not args.skip_styles:
        print("Applying cell styles...", flush=True)
        style_updates = apply_styles(
            spreadsheet_token,
            ws,
            sheet_ref.sheet_id,
            args.identity,
            args.max_style_ranges,
        )

    merge_updates = 0
    merge_warnings: list[str] = []
    if args.copy_merges:
        print("Applying merged cells...", flush=True)
        merge_updates, merge_warnings = apply_merges(spreadsheet_token, ws, sheet_ref.sheet_id, args.identity)

    image_updates = 0
    image_warnings: list[str] = []
    if args.insert_images:
        print("Applying images...", flush=True)
        image_updates, image_warnings = apply_images(spreadsheet_token, ws, sheet_ref.sheet_id, args.identity)

    result = {
        "spreadsheet_token": spreadsheet_token,
        "sheet_id": sheet_ref.sheet_id,
        "source_sheet_title": ws.title,
        "width_updates": width_updates,
        "row_height_updates": row_height_updates,
        "row_height_warnings": row_height_warnings,
        "style_updates": style_updates,
        "merge_updates": merge_updates,
        "merge_warnings": merge_warnings,
        "image_updates": image_updates,
        "image_warnings": image_warnings,
    }
    sys.stdout.flush()
    print(json.dumps(result, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
