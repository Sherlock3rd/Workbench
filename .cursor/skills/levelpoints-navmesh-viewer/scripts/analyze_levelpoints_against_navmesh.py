#!/usr/bin/env python3
import argparse
import csv
import json
import math
import warnings
from collections import Counter, defaultdict, deque
from pathlib import Path


def as_float(value, default=0.0):
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def as_key(value):
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def nested_mb(component):
    return ((component.get("serializedData") or {}).get("MonoBehaviour") or {})


def position_tuple(component):
    pos = component.get("position")
    if not isinstance(pos, dict):
        return None
    return as_float(pos.get("x")), as_float(pos.get("y")), as_float(pos.get("z"))


def rotation_tuple(component):
    rotation = component.get("rotation")
    if not isinstance(rotation, dict):
        return None
    return (
        as_float(rotation.get("x")),
        as_float(rotation.get("y")),
        as_float(rotation.get("z")),
        as_float(rotation.get("w"), 1.0),
    )


def forward_from_quaternion(rotation):
    if rotation is None:
        return None
    x, y, z, w = rotation
    return (
        2 * (x * z + w * y),
        2 * (y * z - w * x),
        1 - 2 * (x * x + y * y),
    )


def polygon_vertices(poly):
    return [(as_float(v.get("x")), as_float(v.get("z"))) for v in poly.get("Vertexs") or []]


def point_in_poly(x, z, verts):
    inside = False
    for i in range(len(verts)):
        x1, z1 = verts[i]
        x2, z2 = verts[(i + 1) % len(verts)]
        if (z1 > z) != (z2 > z):
            x_at_z = (x2 - x1) * (z - z1) / (z2 - z1 + 1e-12) + x1
            if x < x_at_z:
                inside = not inside
    return inside


def build_navmesh(nav):
    polys = []
    for i, poly in enumerate(nav.get("NavMeshPolygons", [])):
        verts = polygon_vertices(poly)
        if len(verts) < 3:
            continue
        xs = [x for x, _ in verts]
        zs = [z for _, z in verts]
        polys.append({
            "source_index": i,
            "polygon_index": poly.get("PolygonIndex", i),
            "province": poly.get("ProvinceID"),
            "group": poly.get("GroupIndex"),
            "area": poly.get("AreaType"),
            "verts": verts,
            "bbox": (min(xs), min(zs), max(xs), max(zs)),
        })
    return polys


def build_spatial_index(polys, cell_size=20.0):
    index = defaultdict(list)
    for i, poly in enumerate(polys):
        x1, z1, x2, z2 = poly["bbox"]
        for cx in range(math.floor(x1 / cell_size), math.floor(x2 / cell_size) + 1):
            for cz in range(math.floor(z1 / cell_size), math.floor(z2 / cell_size) + 1):
                index[(cx, cz)].append(i)
    return index


def candidate_polys(index, x, z, cell_size=20.0):
    cx = math.floor(x / cell_size)
    cz = math.floor(z / cell_size)
    seen = set()
    for dx in (-1, 0, 1):
        for dz in (-1, 0, 1):
            for i in index.get((cx + dx, cz + dz), []):
                if i not in seen:
                    seen.add(i)
                    yield i


def components(polys):
    edge_to_polys = defaultdict(list)
    for i, poly in enumerate(polys):
        verts = poly["verts"]
        for a, b in zip(verts, verts[1:] + verts[:1]):
            key = tuple(sorted(((round(a[0], 3), round(a[1], 3)), (round(b[0], 3), round(b[1], 3)))))
            edge_to_polys[key].append(i)
    graph = [[] for _ in polys]
    for members in edge_to_polys.values():
        if len(members) < 2:
            continue
        for i in members:
            graph[i].extend(j for j in members if j != i)
    comp = [-1] * len(polys)
    sizes = []
    for start in range(len(polys)):
        if comp[start] != -1:
            continue
        cid = len(sizes)
        queue = deque([start])
        comp[start] = cid
        count = 0
        while queue:
            cur = queue.popleft()
            count += 1
            for nxt in graph[cur]:
                if comp[nxt] == -1:
                    comp[nxt] = cid
                    queue.append(nxt)
        sizes.append(count)
    return comp, sizes


def classify_point(point, polys, index, comp, main_comp):
    x, z = point["x"], point["z"]
    for pi in candidate_polys(index, x, z):
        x1, z1, x2, z2 = polys[pi]["bbox"]
        if x < x1 or x > x2 or z < z1 or z > z2:
            continue
        if point_in_poly(x, z, polys[pi]["verts"]):
            cid = comp[pi]
            return {
                "inside": True,
                "poly": polys[pi]["polygon_index"],
                "component": cid,
                "main": cid == main_comp,
                "issue": None if cid == main_comp else "ISOLATED_COMPONENT",
            }
    return {"inside": False, "poly": None, "component": None, "main": False, "issue": "OUTSIDE_MESH"}


def load_labels(path):
    if not path:
        return {}
    label_path = Path(path)
    if not label_path.exists():
        return {}
    return json.loads(label_path.read_text(encoding="utf-8-sig"))


def load_interactive_obj_names(path):
    if not path:
        return {}
    xlsx_path = Path(path)
    if not xlsx_path.exists():
        return {}
    try:
        import openpyxl
    except ImportError:
        return {}
    with warnings.catch_warnings():
        warnings.simplefilter("ignore", UserWarning)
        workbook = openpyxl.load_workbook(xlsx_path, read_only=False, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    headers = [str(sheet.cell(1, col).value or "").strip() for col in range(1, sheet.max_column + 1)]
    id_col = next((i for i, h in enumerate(headers) if h.lower() in ("int_row_id", "row_id", "id")), None)
    name_col = next((i for i, h in enumerate(headers) if h.lower() == "display_name"), None)
    if id_col is None or name_col is None:
        return {}
    names = {}
    for row_index in range(2, sheet.max_row + 1):
        obj_id = as_key(sheet.cell(row_index, id_col + 1).value)
        display_name = sheet.cell(row_index, name_col + 1).value
        if obj_id and display_name not in (None, ""):
            names[obj_id] = str(display_name)
    return names


def label_for(point, labels):
    for key, field in (("objId", "obj_id"), ("objType", "obj_type"), ("className", "class_name")):
        value = as_key(point.get(field))
        mapping = labels.get(key, {})
        if value and value in mapping:
            return mapping[value]
    prefab = (point.get("prefab_path") or "").lower()
    for rule in labels.get("prefabContains", []):
        if str(rule.get("contains", "")).lower() in prefab:
            return rule.get("label")
    return point.get("class_name") or "Unknown"


def object_category_for(point):
    if point.get("npc_id") is not None:
        return "NPC"
    if point.get("team_id") is not None:
        return "Team"
    if point.get("pet_monster_id") is not None:
        return "PetMonster"
    if point.get("obj_type") is not None:
        return f"ObjType {as_key(point.get('obj_type'))}"
    return point.get("class_name") or "Unknown"


def normalize_point(component, index, labels, interactive_obj_names):
    pos = position_tuple(component)
    if pos is None:
        return None
    mb = nested_mb(component)
    obj_id = mb.get("ObjID")
    rotation = rotation_tuple(component)
    forward = forward_from_quaternion(rotation)
    point = {
        "index": index,
        "class_name": component.get("className"),
        "sub_id": component.get("subID") or mb.get("SubID"),
        "parent_sub_id": component.get("parentSubID"),
        "belong_to_layer_id": component.get("belongToLayerID", mb.get("BelongToLayerID")),
        "prefab_path": component.get("prefabPath") or "",
        "obj_id": obj_id,
        "obj_type": mb.get("ObjType"),
        "npc_id": mb.get("NpcID"),
        "pet_monster_id": mb.get("PetMonsterID"),
        "team_id": mb.get("TeamID"),
        "area_id": mb.get("AreaID"),
        "x": pos[0],
        "y": pos[1],
        "z": pos[2],
        "rot_x": rotation[0] if rotation else None,
        "rot_y": rotation[1] if rotation else None,
        "rot_z": rotation[2] if rotation else None,
        "rot_w": rotation[3] if rotation else None,
        "forward_x": forward[0] if forward else None,
        "forward_y": forward[1] if forward else None,
        "forward_z": forward[2] if forward else None,
        "obj_display_name": interactive_obj_names.get(as_key(obj_id)),
    }
    point["label"] = label_for(point, labels)
    point["object_category"] = object_category_for(point)
    return point


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--navmesh", required=True)
    parser.add_argument("--levelpoints", required=True)
    parser.add_argument("--labels")
    parser.add_argument("--interactive-obj-xlsx", default="E:/Beagle/data/GameDatas/datas/InteractiveObj.xlsx")
    parser.add_argument("--out-json", required=True)
    parser.add_argument("--out-csv", required=True)
    args = parser.parse_args()

    nav = json.loads(Path(args.navmesh).read_text(encoding="utf-8-sig"))
    levelpoints = json.loads(Path(args.levelpoints).read_text(encoding="utf-8-sig"))
    labels = load_labels(args.labels)
    interactive_names = load_interactive_obj_names(args.interactive_obj_xlsx)
    polys = build_navmesh(nav)
    comp, comp_sizes = components(polys)
    main_comp = max(range(len(comp_sizes)), key=lambda i: comp_sizes[i]) if comp_sizes else None
    index = build_spatial_index(polys)

    points = []
    skipped = []
    for i, component in enumerate(levelpoints.get("components", [])):
        point = normalize_point(component, i, labels, interactive_names)
        if point is None:
            skipped.append({"index": i, "class_name": component.get("className"), "reason": "NO_POSITION"})
            continue
        point["nav"] = classify_point(point, polys, index, comp, main_comp)
        points.append(point)

    issues = []
    for point in points:
        if point["nav"].get("issue"):
            issues.append({
                "index": point["index"],
                "issue": point["nav"]["issue"],
                "label": point["label"],
                "class_name": point["class_name"],
                "sub_id": point["sub_id"],
                "obj_id": point["obj_id"],
                "obj_display_name": point["obj_display_name"],
                "obj_type": point["obj_type"],
                "npc_id": point["npc_id"],
                "team_id": point["team_id"],
                "area_id": point["area_id"],
                "x": point["x"],
                "y": point["y"],
                "z": point["z"],
                "rot_x": point["rot_x"],
                "rot_y": point["rot_y"],
                "rot_z": point["rot_z"],
                "rot_w": point["rot_w"],
                "forward_x": point["forward_x"],
                "forward_y": point["forward_y"],
                "forward_z": point["forward_z"],
                "component": point["nav"]["component"],
                "poly": point["nav"]["poly"],
                "prefab_path": point["prefab_path"],
            })

    summary = {
        "navmesh": args.navmesh,
        "levelpoints": args.levelpoints,
        "label_config": args.labels,
        "interactive_obj_xlsx": args.interactive_obj_xlsx,
        "level_id": levelpoints.get("levelId"),
        "level_version": levelpoints.get("version"),
        "polygon_count": len(polys),
        "component_count": len(comp_sizes),
        "main_component": main_comp,
        "main_component_size": comp_sizes[main_comp] if main_comp is not None else 0,
        "raw_component_count": len(levelpoints.get("components", [])),
        "point_count": len(points),
        "skipped_component_count": len(skipped),
        "class_counts": Counter(p["class_name"] for p in points),
        "label_counts": Counter(p["label"] for p in points),
        "inside_count": sum(1 for p in points if p["nav"]["inside"]),
        "outside_count": sum(1 for p in points if not p["nav"]["inside"]),
        "isolated_count": sum(1 for p in points if p["nav"]["issue"] == "ISOLATED_COMPONENT"),
        "issue_counts": Counter(i["issue"] for i in issues),
    }

    out_json = Path(args.out_json)
    out_json.parent.mkdir(parents=True, exist_ok=True)
    out_json.write_text(json.dumps({"summary": summary, "issues": issues, "points": points, "skipped": skipped}, ensure_ascii=False, indent=2), encoding="utf-8")

    out_csv = Path(args.out_csv)
    out_csv.parent.mkdir(parents=True, exist_ok=True)
    fields = ["index", "issue", "label", "class_name", "sub_id", "obj_id", "obj_display_name", "obj_type", "npc_id", "team_id", "area_id", "x", "y", "z", "rot_x", "rot_y", "rot_z", "rot_w", "forward_x", "forward_y", "forward_z", "component", "poly", "prefab_path"]
    with out_csv.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fields)
        writer.writeheader()
        writer.writerows(issues)


if __name__ == "__main__":
    main()
