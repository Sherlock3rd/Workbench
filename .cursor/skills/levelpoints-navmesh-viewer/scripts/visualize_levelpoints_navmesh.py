#!/usr/bin/env python3
import argparse
import base64
import json
import subprocess
import sys
from collections import Counter, defaultdict, deque
from pathlib import Path
import warnings


def as_float(value):
    return float(value or 0)


def image_data_uri(path):
    image_path = Path(path)
    if not image_path.exists():
        return ""
    mime = "image/jpeg" if image_path.suffix.lower() in {".jpg", ".jpeg"} else "image/png"
    data = base64.b64encode(image_path.read_bytes()).decode("ascii")
    return f"data:{mime};base64,{data}"


def polygon_area(verts):
    if len(verts) < 3:
        return 0.0
    return abs(sum(x1 * z2 - x2 * z1 for (x1, z1), (x2, z2) in zip(verts, verts[1:] + verts[:1]))) / 2


def polygon_record(poly, fallback_index):
    verts = [[round(as_float(v.get("x")), 4), round(as_float(v.get("z")), 4)] for v in poly.get("Vertexs") or []]
    ys = [as_float(v.get("y")) for v in poly.get("Vertexs") or []]
    xs = [v[0] for v in verts]
    zs = [v[1] for v in verts]
    return {
        "i": poly.get("PolygonIndex", fallback_index),
        "area": poly.get("AreaType"),
        "group": poly.get("GroupIndex"),
        "province": poly.get("ProvinceID"),
        "land": poly.get("LandID"),
        "tile": poly.get("TileIndex"),
        "y": round(sum(ys) / len(ys), 4) if ys else 0,
        "navArea": polygon_area(verts),
        "verts": verts,
        "bbox": [min(xs), min(zs), max(xs), max(zs)] if verts else [0, 0, 0, 0],
    }


def as_int(value, default=0):
    try:
        if value is None or value == "":
            return default
        return int(float(value))
    except (TypeError, ValueError):
        return default


def load_area_config(path):
    if not path:
        return {}
    area_path = Path(path)
    if not area_path.exists():
        return {}
    try:
        import openpyxl
    except ImportError:
        return {}

    with warnings.catch_warnings():
        warnings.simplefilter("ignore", UserWarning)
        workbook = openpyxl.load_workbook(area_path, read_only=False, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    headers = [str(sheet.cell(1, col).value or "").strip() for col in range(1, sheet.max_column + 1)]
    cols = {header: i + 1 for i, header in enumerate(headers)}
    required = ("INT_areaID", "INT_gveType", "INT_pvpType")
    if any(key not in cols for key in required):
        return {}

    result = {}
    for row_index in range(2, sheet.max_row + 1):
        area_id = as_int(sheet.cell(row_index, cols["INT_areaID"]).value, None)
        if area_id is None:
            continue
        gve_type = as_int(sheet.cell(row_index, cols["INT_gveType"]).value, 0)
        pvp_type = as_int(sheet.cell(row_index, cols["INT_pvpType"]).value, 0)
        display_name = sheet.cell(row_index, cols.get("display_name", 0)).value if cols.get("display_name") else None
        if pvp_type:
            area_kind = "PVP区域"
        elif gve_type:
            area_kind = "特殊区域"
        else:
            area_kind = "普通区域"
        result[area_id] = {
            "areaKind": area_kind,
            "areaName": str(display_name) if display_name not in (None, "") else "",
            "gveType": gve_type,
            "pvpType": pvp_type,
        }
    return result


def load_interactive_obj_names(path):
    if not path:
        return {}
    xlsx_path = Path(path)
    if not xlsx_path.exists():
        return {}
    try:
        import openpyxl
    except ImportError:
        return {}

    with warnings.catch_warnings():
        warnings.simplefilter("ignore", UserWarning)
        workbook = openpyxl.load_workbook(xlsx_path, read_only=False, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    headers = [str(sheet.cell(1, col).value or "").strip() for col in range(1, sheet.max_column + 1)]
    id_col = next((i + 1 for i, h in enumerate(headers) if h.lower() == "id"), None)
    name_col = next((i + 1 for i, h in enumerate(headers) if h.lower() == "display_name"), None)
    if not id_col or not name_col:
        return {}

    names = {}
    for row_index in range(2, sheet.max_row + 1):
        obj_id = sheet.cell(row_index, id_col).value
        display_name = sheet.cell(row_index, name_col).value
        if obj_id in (None, "") or display_name in (None, ""):
            continue
        names[str(as_int(obj_id))] = str(display_name)
    return names


def apply_area_config(polygons, area_config):
    default_info = {"areaKind": "未知区域", "areaName": "", "gveType": None, "pvpType": None}
    for poly in polygons:
        info = area_config.get(as_int(poly.get("province"), None), default_info)
        poly.update(info)


def components(polygons):
    edge_to_polys = defaultdict(list)
    for i, poly in enumerate(polygons):
        verts = [tuple(v) for v in poly["verts"]]
        for a, b in zip(verts, verts[1:] + verts[:1]):
            key = tuple(sorted(((round(a[0], 3), round(a[1], 3)), (round(b[0], 3), round(b[1], 3)))))
            edge_to_polys[key].append(i)
    graph = [[] for _ in polygons]
    for members in edge_to_polys.values():
        if len(members) < 2:
            continue
        for i in members:
            graph[i].extend(j for j in members if j != i)
    comp = [-1] * len(polygons)
    sizes = []
    for start in range(len(polygons)):
        if comp[start] != -1:
            continue
        cid = len(sizes)
        q = deque([start])
        comp[start] = cid
        count = 0
        while q:
            cur = q.popleft()
            count += 1
            for nxt in graph[cur]:
                if comp[nxt] == -1:
                    comp[nxt] = cid
                    q.append(nxt)
        sizes.append(count)
    return comp, sizes


def boundary_segments(polygons):
    edges = defaultdict(list)
    for poly in polygons:
        verts = poly["verts"]
        for a, b in zip(verts, verts[1:] + verts[:1]):
            key = tuple(sorted(((round(a[0], 3), round(a[1], 3)), (round(b[0], 3), round(b[1], 3)))))
            edges[key].append((a, b, poly))
    segments = []
    for members in edges.values():
        signatures = {region_key(m[2]) for m in members}
        if len(members) != 1 and len(signatures) <= 1:
            continue
        a, b, poly = members[0]
        segments.append({
            "a": a,
            "b": b,
            "bbox": [min(a[0], b[0]), min(a[1], b[1]), max(a[0], b[0]), max(a[1], b[1])],
            "kind": "outer" if len(members) == 1 else "partition",
            "area": poly.get("area"),
            "group": poly.get("group"),
            "province": poly.get("province"),
            "component": poly.get("component"),
            "mainComponent": poly.get("mainComponent"),
            "y": poly.get("y"),
            "vertexCount": len(poly.get("verts") or []),
            "regionKeys": sorted(signatures),
            "areaKind": poly.get("areaKind"),
            "areaName": poly.get("areaName"),
            "gveType": poly.get("gveType"),
            "pvpType": poly.get("pvpType"),
        })
    return segments


def region_key(poly):
    return "|".join(str(poly.get(key)) for key in ("province", "group", "component", "area"))


def assign_boundary_component_areas(segments):
    side_endpoint_to_segments = defaultdict(list)
    outer_endpoint_to_segments = defaultdict(list)
    for i, segment in enumerate(segments):
        if segment.get("kind") == "outer":
            a = tuple(round(v, 3) for v in segment["a"])
            b = tuple(round(v, 3) for v in segment["b"])
            outer_endpoint_to_segments[a].append(i)
            outer_endpoint_to_segments[b].append(i)
            continue
        for region in segment.get("regionKeys") or []:
            a = tuple(round(v, 3) for v in segment["a"])
            b = tuple(round(v, 3) for v in segment["b"])
            side_endpoint_to_segments[(region, a)].append(i)
            side_endpoint_to_segments[(region, b)].append(i)

    visited_outer = set()
    for start in range(len(segments)):
        if segments[start].get("kind") != "outer" or start in visited_outer:
            continue
        stack = [start]
        members = []
        visited_outer.add(start)
        while stack:
            cur = stack.pop()
            members.append(cur)
            for endpoint in (tuple(round(v, 3) for v in segments[cur]["a"]), tuple(round(v, 3) for v in segments[cur]["b"])):
                for nxt in outer_endpoint_to_segments[endpoint]:
                    if nxt not in visited_outer:
                        visited_outer.add(nxt)
                        stack.append(nxt)

        xs = []
        zs = []
        for idx in members:
            xs.extend([segments[idx]["a"][0], segments[idx]["b"][0]])
            zs.extend([segments[idx]["a"][1], segments[idx]["b"][1]])
        bbox_area = (max(xs) - min(xs)) * (max(zs) - min(zs)) if xs and zs else 0.0
        for idx in members:
            segments[idx]["boundaryComponentArea"] = bbox_area
            segments[idx]["boundaryComponentSize"] = len(members)

    visited = set()
    for start in range(len(segments)):
        if segments[start].get("kind") == "outer":
            continue
        for region in segments[start].get("regionKeys") or []:
            state = (region, start)
            if state in visited:
                continue
            stack = [start]
            members = []
            visited.add(state)
            while stack:
                cur = stack.pop()
                members.append(cur)
                segment = segments[cur]
                for endpoint in (tuple(round(v, 3) for v in segment["a"]), tuple(round(v, 3) for v in segment["b"])):
                    for nxt in side_endpoint_to_segments[(region, endpoint)]:
                        nxt_state = (region, nxt)
                        if nxt_state not in visited:
                            visited.add(nxt_state)
                            stack.append(nxt)

            xs = []
            zs = []
            for idx in members:
                xs.extend([segments[idx]["a"][0], segments[idx]["b"][0]])
                zs.extend([segments[idx]["a"][1], segments[idx]["b"][1]])
            bbox_area = (max(xs) - min(xs)) * (max(zs) - min(zs)) if xs and zs else 0.0
            for idx in members:
                current = segments[idx].get("boundaryComponentArea")
                segments[idx]["boundaryComponentArea"] = bbox_area if current is None else min(current, bbox_area)
                segments[idx]["boundaryComponentSize"] = max(segments[idx].get("boundaryComponentSize", 0), len(members))
    return segments


def add_navmesh_meta(polygons):
    comp, sizes = components(polygons)
    main_comp = max(range(len(sizes)), key=lambda i: sizes[i]) if sizes else None
    for i, poly in enumerate(polygons):
        poly["component"] = comp[i]
        poly["componentSize"] = sizes[comp[i]]
        poly["mainComponent"] = comp[i] == main_comp
    return main_comp, sizes


def stats_for(nav, polygons, source, analysis_path, levelpoints_source):
    xs = [x for poly in polygons for x, _ in poly["verts"]]
    zs = [z for poly in polygons for _, z in poly["verts"]]
    ys = [poly["y"] for poly in polygons] or [0]
    vertex_counts = [len(poly["verts"]) for poly in polygons] or [0]
    return {
        "source": str(source),
        "analysis": str(analysis_path),
        "levelpointsSource": str(levelpoints_source),
        "mapWidth": nav.get("MapWidth"),
        "mapHeight": nav.get("MapHeight"),
        "startX": nav.get("StartX"),
        "startZ": nav.get("StartZ"),
        "polygonCount": len(polygons),
        "vertexCount": sum(vertex_counts),
        "vertexMin": min(vertex_counts),
        "vertexMax": max(vertex_counts),
        "vertexAvg": round(sum(vertex_counts) / len(vertex_counts), 2) if vertex_counts else 0,
        "bounds": {
            "minX": min(xs) if xs else 0,
            "maxX": max(xs) if xs else 1,
            "minZ": min(zs) if zs else 0,
            "maxZ": max(zs) if zs else 1,
            "minY": min(ys),
            "maxY": max(ys),
        },
    }


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("navmesh")
    parser.add_argument("analysis")
    parser.add_argument("output")
    parser.add_argument("--area-config", default="E:/Beagle/data/GameDatas/datas/area_config.xlsx")
    parser.add_argument("--levelpoints", help="Optional raw LevelPoints.json. When provided, regenerate analysis before writing HTML.")
    parser.add_argument("--labels")
    parser.add_argument("--interactive-obj-xlsx", default="E:/Beagle/data/GameDatas/datas/InteractiveObj.xlsx")
    parser.add_argument("--issues-csv")
    args = parser.parse_args()

    navmesh_path = Path(args.navmesh)
    analysis_path = Path(args.analysis)
    if args.levelpoints:
        analyzer = Path(__file__).with_name("analyze_levelpoints_against_navmesh.py")
        issues_csv = Path(args.issues_csv) if args.issues_csv else analysis_path.with_name(analysis_path.stem.replace("_analysis", "_issues") + ".csv")
        cmd = [
            sys.executable,
            str(analyzer),
            "--navmesh",
            str(navmesh_path),
            "--levelpoints",
            str(Path(args.levelpoints)),
            "--interactive-obj-xlsx",
            str(Path(args.interactive_obj_xlsx)),
            "--out-json",
            str(analysis_path),
            "--out-csv",
            str(issues_csv),
        ]
        if args.labels:
            cmd.extend(["--labels", str(Path(args.labels))])
        subprocess.run(cmd, check=True)

    nav = json.loads(navmesh_path.read_text(encoding="utf-8-sig"))
    analysis = json.loads(analysis_path.read_text(encoding="utf-8-sig"))
    polygons = [polygon_record(poly, i) for i, poly in enumerate(nav.get("NavMeshPolygons", []))]
    polygons = [poly for poly in polygons if len(poly["verts"]) >= 3]
    area_config = load_area_config(args.area_config)
    apply_area_config(polygons, area_config)
    main_comp, comp_sizes = add_navmesh_meta(polygons)
    segments = assign_boundary_component_areas(boundary_segments(polygons))
    stats = stats_for(nav, polygons, navmesh_path, analysis_path, analysis.get("summary", {}).get("levelpoints"))
    stats["componentCount"] = len(comp_sizes)
    stats["mainComponent"] = main_comp
    stats["mainComponentSize"] = comp_sizes[main_comp] if main_comp is not None else 0
    stats["boundarySegmentCount"] = len(segments)
    stats["minimalBoundaryAreaThreshold"] = 200
    stats["areaConfig"] = str(args.area_config)

    interactive_obj_names = load_interactive_obj_names(args.interactive_obj_xlsx)
    payload = json.dumps({"stats": stats, "polygons": polygons, "boundarySegments": segments, "analysis": analysis, "areaConfig": area_config, "interactiveObjNames": interactive_obj_names}, ensure_ascii=False, separators=(",", ":"))
    top_banner_uri = image_data_uri(Path(__file__).with_name("viewer-top-banner.png"))

    html = r"""<!doctype html>
<html lang="zh-CN">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>LevelPoints NavMesh Viewer</title>
<style>
:root{--bg:#f5f7fa;--panel:#fff;--text:#172033;--muted:#607084;--line:#d8dee8}
*{box-sizing:border-box}html,body{height:100%;overflow:hidden}body{margin:0;background:var(--bg);color:var(--text);font-family:"Segoe UI",Arial,sans-serif}
.app{height:100vh;min-height:720px;display:grid;grid-template-columns:370px minmax(0,1fr);overflow:hidden}aside{height:100vh;background:var(--panel);border-right:1px solid var(--line);overflow-y:auto;overflow-x:hidden;padding:18px}
main{min-width:0;min-height:0;display:flex;flex-direction:column;overflow:hidden}h1{margin:0 0 8px;font-size:19px}h2{margin:18px 0 8px;font-size:13px;color:var(--muted);text-transform:uppercase;letter-spacing:.04em}
p{margin:8px 0;color:var(--muted);font-size:13px;line-height:1.5;overflow-wrap:anywhere}.metric{display:grid;grid-template-columns:1fr auto;gap:8px;padding:7px 0;border-bottom:1px solid #edf0f4;font-size:13px}
.toolbar{position:sticky;top:0;z-index:20;flex:0 0 auto;min-height:74px;display:flex;flex-wrap:wrap;align-items:center;gap:8px;padding:8px clamp(190px,24vw,360px) 8px 12px;border-bottom:1px solid var(--line);background:linear-gradient(90deg,rgba(255,255,255,.98),rgba(255,255,255,.92));box-shadow:0 2px 8px rgba(15,23,42,.06);overflow:visible}
.toolbar::after{content:"";position:absolute;right:8px;top:50%;width:clamp(180px,22vw,340px);height:calc(100% - 10px);transform:translateY(-50%);background:linear-gradient(90deg,rgba(255,255,255,.98),rgba(255,255,255,.42) 24%,rgba(255,255,255,0) 48%),url("__TOP_BANNER__") right center/contain no-repeat;pointer-events:none}
button,select,input{border:1px solid var(--line);border-radius:6px;background:#fff;color:var(--text);padding:6px 9px;font-size:13px}button{cursor:pointer}button:hover{background:#eef2f7}
label{display:inline-flex;align-items:center;gap:6px;font-size:13px;color:var(--muted)}input[type=search]{width:260px}.canvas-wrap{position:relative;flex:1 1 auto;min-height:0;background:#e9eef5;overflow:hidden;overscroll-behavior:none}
.province-filter{position:relative;z-index:40}.province-button{min-width:128px;text-align:left}.province-panel{position:absolute;top:calc(100% + 8px);left:0;width:280px;max-height:380px;display:none;background:#fff;border:1px solid var(--line);border-radius:10px;box-shadow:0 18px 48px rgba(15,23,42,.18);padding:10px}.province-filter.open .province-panel{display:block}.province-panel input[type=search]{width:100%;margin-bottom:8px}.province-actions{display:flex;gap:8px;margin-bottom:8px}.province-actions button{flex:1}.province-list{display:grid;gap:4px;max-height:250px;overflow:auto;border:1px solid #edf0f4;border-radius:8px;padding:6px}.province-row{display:grid;grid-template-columns:auto 1fr auto;gap:7px;align-items:center;padding:4px 3px;color:var(--muted);font-size:12px}.province-count{margin:8px 0 0;color:var(--muted);font-size:12px}
canvas{width:100%;height:100%;display:block;cursor:grab}body.connectivity-mode canvas{cursor:url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' width='32' height='32' viewBox='0 0 32 32'%3E%3Cpath fill='%232563eb' stroke='%231e293b' stroke-width='1.5' d='M8 6l8 8-7 7-6-6 7-7-2-2z'/%3E%3Cpath fill='%2322c55e' stroke='%23166534' stroke-width='1.2' d='M17 15l7 7c1 1 1 3 0 4s-3 1-4 0l-7-7z'/%3E%3Ccircle cx='25' cy='10' r='3' fill='%23ef4444'/%3E%3C/svg%3E") 5 24,crosshair}canvas.dragging{cursor:grabbing}body.connectivity-mode canvas.dragging{cursor:grabbing}#tip{position:absolute;pointer-events:none;display:none;background:rgba(255,255,255,.96);border:1px solid var(--line);border-radius:6px;padding:8px 9px;font-size:12px;box-shadow:0 6px 18px rgba(20,32,48,.14);min-width:250px;max-width:390px;overflow-wrap:anywhere}
.filter-tools{display:grid;gap:7px;margin-bottom:8px}.filter-tools input[type=search]{width:100%}.filter-actions{display:flex;gap:7px}.filter-actions button{flex:1}.filters{display:grid;gap:8px;max-height:300px;overflow:auto;border:1px solid #edf0f4;border-radius:6px;padding:8px}.filter-category{border:1px solid #e5eaf2;border-radius:7px;overflow:hidden;background:#fff}.filter-category-title{display:grid;grid-template-columns:auto 1fr auto;gap:7px;align-items:center;width:100%;border:0;border-radius:0;background:#f8fafc;text-align:left;font-weight:650;color:#334155}.filter-category-title:hover{background:#eef2f7}.filter-category-count{font-size:12px;color:var(--muted);font-weight:650}.filter-category-body{display:grid;gap:5px;padding:6px}.filter-row{display:grid;grid-template-columns:auto 1fr auto;gap:7px;align-items:center;font-size:12px;color:var(--muted)}
.swatch{width:11px;height:11px;border-radius:50%;border:1px solid rgba(0,0,0,.15)}table{width:100%;border-collapse:collapse;font-size:12px}td{border-bottom:1px solid #edf0f4;padding:5px 0}td:last-child{text-align:right;font-weight:650}
.legend-title{display:flex;align-items:center;justify-content:space-between;gap:8px;margin:18px 0 8px}.legend-title h2{margin:0}.legend-toggle{font-size:12px;padding:4px 7px}.legend-table.collapsed tr:nth-child(n+6){display:none}
.note,.source-panel{background:#eef5ff;border:1px solid #bdd4ff;color:#24446f;border-radius:6px;padding:10px;font-size:13px;line-height:1.45}.status{color:var(--muted);font-size:12px;margin-left:auto}
.source-row{display:grid;grid-template-columns:72px 1fr;gap:8px;align-items:center;margin:6px 0}.source-row input{width:100%;font-size:12px}.button-row{display:flex;flex-wrap:wrap;gap:7px}.mini{font-size:12px;padding:5px 7px}
.check-options{display:grid;gap:7px;margin:8px 0}.progress{height:10px;background:#d9e7fb;border-radius:999px;overflow:hidden;margin:8px 0}.progress>div{height:100%;width:0;background:#2563eb;transition:width .18s ease}
.error-list{display:grid;gap:7px;max-height:260px;overflow:auto;margin-top:8px}.error-item{border:1px solid #f2b8b5;background:#fff7f7;border-radius:6px;padding:7px;text-align:left;font-size:12px;color:#5f1b1b}.error-item:hover{background:#ffecec}.error-item b{display:block;color:#b42318;margin-bottom:3px}.error-item span{display:block;color:#7a3a36;overflow-wrap:anywhere}
.distance-item{border:1px solid #f7c873;background:#fffaf0;border-radius:6px;padding:7px;text-align:left;font-size:12px;color:#63420d}.distance-item:hover{background:#fff3d8}.distance-item b{display:block;color:#9a3412;margin-bottom:3px}.distance-item span{display:block;color:#7c4a03;overflow-wrap:anywhere}
@media(max-width:920px){.app{grid-template-columns:1fr;grid-template-rows:auto 1fr}aside{max-height:44vh;border-right:0;border-bottom:1px solid var(--line)}}
</style>
</head>
<body>
<div class="app">
<aside>
<h1>LevelPoints NavMesh Viewer</h1>
<p>Navmesh 投影到 X/Z 平面，点位来自 Unity LevelEditor <code>components[]</code>。</p>
<div class="note">点位默认按 <code>ObjID</code> 分类和着色；悬停和定位面板会显示点位在原始 <code>LevelPoints.json</code> 里的 <code>components[]</code> 索引。</div>
<h2>概览</h2><div id="metrics"></div>
<h2>点位筛选</h2><p>点位固定按 <code>ObjID</code> 分类，可按大分类搜索、全选或全不选。</p>
<div class="filter-tools">
  <input id="filter-search" type="search" placeholder="搜索 ObjID / NPC / Team / PetMonster">
  <div class="filter-actions"><button id="filter-all" class="mini" type="button">全选</button><button id="filter-none" class="mini" type="button">全不选</button></div>
</div>
<div id="filters" class="filters"></div>
<div class="legend-title"><h2>点位分类颜色</h2><button id="raw-type-toggle" class="mini legend-toggle" type="button">展开全部</button></div><table id="raw-type-table" class="legend-table collapsed"></table>
<h2>合法性检查</h2>
<div class="source-panel">
  <div class="check-options">
    <label><input id="check-navmesh-overlap" type="checkbox" checked> 点位是否落在 NavMesh 上</label>
    <label><input id="check-station-validity" type="checkbox" checked> 传送车站合法性检查（PVP + 前方5m NavMesh）</label>
  </div>
  <div class="button-row"><button id="run-validation" class="mini">启动合法性检查</button><button id="clear-validation" class="mini">清空检查结果</button></div>
  <div class="progress"><div id="validation-progress-bar"></div></div>
  <p id="validation-status">未开始检查。</p>
  <div id="validation-results" class="error-list"></div>
</div>
<h2>车站距离检查</h2>
<div class="source-panel">
  <p class="hint">计算始终基于完整 NavMesh polygon 共边图，不受当前底图简化显示模式影响。</p>
  <div class="button-row"><button id="run-station-distance" class="mini">启动车站距离检查</button><button id="clear-station-distance" class="mini">清空车站连线</button></div>
  <div class="progress"><div id="station-distance-progress-bar"></div></div>
  <p id="station-distance-status">未开始检查。</p>
  <div id="station-distance-results" class="error-list"></div>
</div>
<h2>源文件路径 / 重新加载</h2>
<div class="source-panel">
  <div class="source-row"><span>NavMesh</span><input id="navmesh-path" type="text" readonly></div>
  <div class="source-row"><span>LevelPoints</span><input id="levelpoints-path" type="text" readonly></div>
  <div class="button-row">
    <button id="choose-navmesh" class="mini">更改 NavMesh 文件</button>
    <button id="choose-levelpoints" class="mini">更改 LevelPoints 文件</button>
    <button id="refresh-source" class="mini">刷新重新加载</button>
  </div>
  <p id="source-status">点击“刷新重新加载”会重新读取已授权的源文件并重跑分析；Electron 桌面版会由主进程按当前路径直接读取磁盘最新文件。</p>
  <div class="progress"><div id="source-progress-bar"></div></div>
  <p id="source-progress-status">未开始加载。</p>
  <input id="navmesh-file" type="file" accept=".json,application/json" hidden>
  <input id="levelpoints-file" type="file" accept=".json,application/json" hidden>
</div>
<h2>Top Label</h2><table id="label-table"></table>
<h2>Top Class</h2><table id="class-table"></table>
</aside>
<main>
<div class="toolbar">
<button id="zoom-in">放大</button><button id="zoom-out">缩小</button><button id="reset">重置视图</button>
<label>底图 <select id="base-mode"><option value="minimal" selected>极简模式</option><option value="full">完整 NavMesh</option><option value="outline">分区边界简化</option></select></label>
<label>NavMesh 着色 <select id="color-mode"><option value="province">ProvinceID</option><option value="area">区域类型</option><option value="height">Height Y</option></select></label>
<label>整体旋转 <select id="base-rotation"><option value="0" selected>0°</option><option value="90">90°</option><option value="180">180°</option><option value="270">270°</option></select></label>
<label>整体镜像 <select id="view-mirror"><option value="none">无</option><option value="x">左右镜像</option><option value="z" selected>上下镜像</option><option value="xz">左右+上下</option></select></label>
<button id="connectivity-toggle" type="button">连通性检查</button>
<div id="province-filter" class="province-filter">
  <button id="province-filter-button" class="province-button" type="button">省份：全部</button>
  <div class="province-panel">
    <input id="province-search" type="search" placeholder="搜索 ProvinceID">
    <div class="province-actions"><button id="province-all" type="button">全选</button><button id="province-none" type="button">全不选</button></div>
    <div id="province-list" class="province-list"></div>
    <p id="province-count" class="province-count">已选择全部省份</p>
  </div>
</div>
<input id="search" type="search" placeholder="搜索点位 UID / subID">
<label><input id="stroke-toggle" type="checkbox" checked>边线</label><label><input id="points-toggle" type="checkbox" checked>点位</label><label><input id="issues-toggle" type="checkbox">只看问题</label><label><input id="hover-toggle" type="checkbox" checked>悬停信息</label>
<span id="status" class="status"></span>
</div>
<div class="canvas-wrap"><canvas id="map"></canvas><div id="tip"></div></div>
</main>
</div>
<script>
window.__viewerErrors=[];
window.addEventListener('error',e=>window.__viewerErrors.push(`${e.message} @ ${e.filename}:${e.lineno}:${e.colno}`));
window.addEventListener('unhandledrejection',e=>window.__viewerErrors.push(`unhandled rejection: ${e.reason&&e.reason.stack?e.reason.stack:e.reason}`));
const DATA=__PAYLOAD__;
let stats=DATA.stats,polygons=DATA.polygons||[],boundarySegments=DATA.boundarySegments||[],analysis=DATA.analysis||{},points=analysis.points||[],areaConfig=DATA.areaConfig||{},interactiveObjNames=DATA.interactiveObjNames||{};
const canvas=document.getElementById('map'),ctx=canvas.getContext('2d'),tip=document.getElementById('tip');
const baseModeEl=document.getElementById('base-mode'),modeEl=document.getElementById('color-mode'),rotationEl=document.getElementById('base-rotation'),mirrorEl=document.getElementById('view-mirror'),searchEl=document.getElementById('search');
const strokeEl=document.getElementById('stroke-toggle'),pointsEl=document.getElementById('points-toggle'),issuesEl=document.getElementById('issues-toggle'),hoverEl=document.getElementById('hover-toggle');
const statusEl=document.getElementById('status'),filtersEl=document.getElementById('filters'),sourceStatusEl=document.getElementById('source-status');
const validationStatusEl=document.getElementById('validation-status'),validationResultsEl=document.getElementById('validation-results'),validationProgressBarEl=document.getElementById('validation-progress-bar');
const stationDistanceStatusEl=document.getElementById('station-distance-status'),stationDistanceResultsEl=document.getElementById('station-distance-results'),stationDistanceProgressBarEl=document.getElementById('station-distance-progress-bar');
const sourceProgressBarEl=document.getElementById('source-progress-bar'),sourceProgressStatusEl=document.getElementById('source-progress-status');
const connectivityToggleEl=document.getElementById('connectivity-toggle');
const navmeshFileEl=document.getElementById('navmesh-file'),levelpointsFileEl=document.getElementById('levelpoints-file'),navmeshPathEl=document.getElementById('navmesh-path'),levelpointsPathEl=document.getElementById('levelpoints-path');
const rawTypeTableEl=document.getElementById('raw-type-table'),rawTypeToggleEl=document.getElementById('raw-type-toggle');
const filterSearchEl=document.getElementById('filter-search'),filterAllEl=document.getElementById('filter-all'),filterNoneEl=document.getElementById('filter-none');
const provinceFilterEl=document.getElementById('province-filter'),provinceButtonEl=document.getElementById('province-filter-button'),provinceSearchEl=document.getElementById('province-search'),provinceListEl=document.getElementById('province-list'),provinceCountEl=document.getElementById('province-count');
const palette=['#2878bd','#36a168','#e6a01b','#cc4c4c','#7b61c9','#2aa6b1','#d45b9f','#6c8a1f','#b7672c','#526a86','#009a78','#b3446c','#2563eb','#dc2626','#16a34a','#9333ea'];
const minimalBoundaryAreaThreshold=stats.minimalBoundaryAreaThreshold||200;
let dpr=Math.max(1,Math.min(2,window.devicePixelRatio||1)),view={},dragging=false,dragMoved=false,last=null,enabledGroups=new Set(),filterEntries=[],collapsedFilterCategories=new Set(),filterQuery='',selectedPoint=null,polyIndex=null,boundaryIndex=null,polygonById=new Map(),navGraph=null,connectivityGraph=null,connectivityMode=false,connectivityStartPoly=null,connectedPolyIds=new Set(),lastNavmeshFile=null,lastLevelpointsFile=null,lastNavmeshHandle=null,lastLevelpointsHandle=null,validationErrors=[],validationRunning=false,stationDistanceLines=[],stationDistanceTopIds=new Set(),stationDistanceRunning=false,allProvinceIds=[],selectedProvinces=new Set(),provinceQuery='';

function fmt(v,n=2){return v===null||v===undefined?'-':typeof v==='number'?v.toFixed(n):String(v)}
function valueText(v){return v===null||v===undefined||v===''?'(empty)':String(v)}
function hashColor(value){const s=String(value);let h=0;for(let i=0;i<s.length;i++)h=((h<<5)-h+s.charCodeAt(i))|0;return palette[Math.abs(h)%palette.length]}
function minMaxValues(items,getter){let min=Infinity,max=-Infinity;for(const item of items){const v=getter?getter(item):item;if(!Number.isFinite(v))continue;if(v<min)min=v;if(v>max)max=v}return{min:min===Infinity?0:min,max:max===-Infinity?0:max}}
function bboxForPoints(items,getX,getZ){let minX=Infinity,minZ=Infinity,maxX=-Infinity,maxZ=-Infinity;for(const item of items){const x=getX(item),z=getZ(item);if(Number.isFinite(x)){if(x<minX)minX=x;if(x>maxX)maxX=x}if(Number.isFinite(z)){if(z<minZ)minZ=z;if(z>maxZ)maxZ=z}}return{minX:minX===Infinity?0:minX,minZ:minZ===Infinity?0:minZ,maxX:maxX===-Infinity?0:maxX,maxZ:maxZ===-Infinity?0:maxZ}}
function rawTypeText(p){if(p.npc_id!==null&&p.npc_id!==undefined)return `NPC ${p.npc_id}`;if(p.team_id!==null&&p.team_id!==undefined)return `Team ${p.team_id}`;if(p.pet_monster_id!==null&&p.pet_monster_id!==undefined)return `PetMonster ${p.pet_monster_id}`;if(p.obj_type!==null&&p.obj_type!==undefined){const name=p.obj_display_name?` / ${p.obj_display_name}`:'';return `ObjType ${p.obj_type} / ObjID ${valueText(p.obj_id)}${name}`}return valueText(p.class_name)}
function objectIdText(p){if(p.obj_id!==null&&p.obj_id!==undefined)return `ObjID ${valueText(p.obj_id)}`;return rawTypeText(p)}
function groupValue(p){return objectIdText(p)}
function interactiveObjName(objId){if(objId===null||objId===undefined)return null;return interactiveObjNames[valueText(objId)]||interactiveObjNames[String(Number(objId))]||null}
function pointDisplayText(p){if(p.obj_id!==null&&p.obj_id!==undefined){const name=p.obj_display_name||interactiveObjName(p.obj_id);return `${objectIdText(p)}${name?` / ${name}`:''}`}return rawTypeText(p)}
function pointFilterCategory(p){if(p.obj_id!==null&&p.obj_id!==undefined)return'Obj';if(p.pet_monster_id!==null&&p.pet_monster_id!==undefined)return'PetMonster';if(p.team_id!==null&&p.team_id!==undefined)return'Team';if(p.npc_id!==null&&p.npc_id!==undefined)return'NPC';return'Other'}
function pointColor(p){return hashColor(groupValue(p))}
function colorFor(poly){const m=modeEl.value;if(m==='province')return hashColor(poly.province);if(m==='area'){if(poly.areaKind==='PVP区域')return '#dc2626';if(poly.areaKind==='特殊区域')return '#9333ea';if(poly.areaKind==='普通区域')return '#36a168';return '#64748b'}if(m==='height')return heightColor(poly.y);return hashColor(poly.province)}
function heightColor(y){const b=stats.bounds,t=b.maxY===b.minY?0:(y-b.minY)/(b.maxY-b.minY);return `rgb(${Math.round(34+t*210)},${Math.round(110+(1-Math.abs(t-.45)*1.5)*85)},${Math.round(190-t*150)})`}
function buildSpatialIndex(items,bboxFn,cellSize){const cells=new Map();for(let i=0;i<items.length;i++){const b=bboxFn(items[i]);for(let x=Math.floor(b[0]/cellSize);x<=Math.floor(b[2]/cellSize);x++)for(let z=Math.floor(b[1]/cellSize);z<=Math.floor(b[3]/cellSize);z++){const k=`${x},${z}`;if(!cells.has(k))cells.set(k,[]);cells.get(k).push(i)}}return{items,cells,cellSize}}
function querySpatialIndex(index,bbox){const out=[],seen=new Set();for(let x=Math.floor(bbox[0]/index.cellSize);x<=Math.floor(bbox[2]/index.cellSize);x++)for(let z=Math.floor(bbox[1]/index.cellSize);z<=Math.floor(bbox[3]/index.cellSize);z++){const bucket=index.cells.get(`${x},${z}`)||[];for(const i of bucket){if(seen.has(i))continue;seen.add(i);const item=index.items[i],b=item.bbox;if(b[2]>=bbox[0]&&b[0]<=bbox[2]&&b[3]>=bbox[1]&&b[1]<=bbox[3])out.push(item)}}return out}
function rebuildIndexes(){polyIndex=buildSpatialIndex(polygons,p=>p.bbox,80);boundaryIndex=buildSpatialIndex(boundarySegments,s=>s.bbox,120);polygonById=new Map(polygons.map(p=>[String(p.i),p]))}
function edgeKey(a,b){return [`${a[0].toFixed(3)},${a[1].toFixed(3)}`,`${b[0].toFixed(3)},${b[1].toFixed(3)}`].sort().join('|')}
function sharedPortalLength(a,b){return Math.hypot(a[0]-b[0],a[1]-b[1])}
function buildConnectivityGraph(){if(connectivityGraph)return connectivityGraph;const adj=new Map(),edges=new Map();for(const poly of polygons){adj.set(String(poly.i),new Set());const vs=poly.verts;for(let i=0;i<vs.length;i++){const a=vs[i],b=vs[(i+1)%vs.length],key=edgeKey(a,b);if(!edges.has(key))edges.set(key,[]);edges.get(key).push({poly,a,b})}}for(const members of edges.values()){if(members.length<2)continue;const portal=sharedPortalLength(members[0].a,members[0].b);if(portal<2)continue;for(let i=0;i<members.length;i++)for(let j=i+1;j<members.length;j++){const a=String(members[i].poly.i),b=String(members[j].poly.i);adj.get(a).add(b);adj.get(b).add(a)}}connectivityGraph=adj;return adj}
function clearConnectivityCheck(){connectivityMode=false;connectivityStartPoly=null;connectedPolyIds=new Set();document.body.classList.remove('connectivity-mode');connectivityToggleEl.textContent='连通性检查';connectivityToggleEl.style.background='';tip.style.display='none';statusEl.textContent='已取消连通性检查，点位显示恢复为原筛选状态。';draw()}
function toggleConnectivityCheck(){if(connectivityMode){clearConnectivityCheck();return}connectivityMode=true;connectivityStartPoly=null;connectedPolyIds=new Set();selectedPoint=null;document.body.classList.add('connectivity-mode');connectivityToggleEl.textContent='取消连通性检查';connectivityToggleEl.style.background='#dcfce7';tip.style.display='none';statusEl.textContent='连通性检查：点击任意 NavMesh 区域开始检查，点位会临时隐藏。';draw()}
function runConnectivityCheck(poly){if(!poly)return;const graph=buildConnectivityGraph(),start=String(poly.i),seen=new Set([start]),stack=[start];while(stack.length){const cur=stack.pop();for(const next of graph.get(cur)||[]){if(!seen.has(next)){seen.add(next);stack.push(next)}}}connectivityStartPoly=start;connectedPolyIds=seen;statusEl.textContent=`连通性检查：起点 Polygon ${start}，连通 ${seen.size}/${polygons.length} 个 polygon。`;draw()}
function connectivityColor(poly){if(!connectivityMode||!connectivityStartPoly)return null;return connectedPolyIds.has(String(poly.i))?'connected':'blocked'}
function rebuildProvinceFilter(reset=true){allProvinceIds=Array.from(new Set(polygons.map(p=>valueText(p.province)))).sort((a,b)=>Number(a)-Number(b)||a.localeCompare(b));if(reset)selectedProvinces=new Set(allProvinceIds);renderProvinceFilter()}
function provinceSelected(poly){return allProvinceIds.length===0||selectedProvinces.has(valueText(poly.province))}
function matchedPolygonForPoint(p){if(!p.nav||!p.nav.inside)return null;return polygonById.get(String(p.nav.poly))||null}
function pointProvinceSelected(p){const poly=matchedPolygonForPoint(p);return !!poly&&provinceSelected(poly)}
function provinceSelectionText(){if(!allProvinceIds.length)return'省份：无';if(selectedProvinces.size===allProvinceIds.length)return`省份：全部(${allProvinceIds.length})`;return`省份：${selectedProvinces.size}/${allProvinceIds.length}`}
function renderProvinceFilter(){provinceButtonEl.textContent=provinceSelectionText();provinceCountEl.textContent=allProvinceIds.length?`已选择 ${selectedProvinces.size} / ${allProvinceIds.length} 个省份`:'没有可筛选省份';const q=provinceQuery.trim().toLowerCase(),counts=countBy(polygons,p=>p.province);const rows=allProvinceIds.filter(id=>!q||id.toLowerCase().includes(q));provinceListEl.innerHTML=rows.map(id=>`<label class="province-row"><input type="checkbox" data-province="${escapeHtml(id)}" ${selectedProvinces.has(id)?'checked':''}><span>ProvinceID ${escapeHtml(id)}</span><b>${counts.get(id)||0}</b></label>`).join('')||'<p class="province-count">没有匹配省份</p>';provinceListEl.querySelectorAll('input').forEach(input=>input.addEventListener('change',()=>{input.checked?selectedProvinces.add(input.dataset.province):selectedProvinces.delete(input.dataset.province);renderProvinceFilter();fillStats();draw()}))}
function displayBounds(){const b=stats.bounds,corners=[[b.minX,b.minZ],[b.maxX,b.minZ],[b.minX,b.maxZ],[b.maxX,b.maxZ]].map(([x,z])=>rawToDisplay(x,z));return bboxForPoints(corners,p=>p.x,p=>p.z)}
function fitViewToCanvas(centerX,centerZ,width,height){const canvasAspect=canvas.height/Math.max(canvas.width,1),safeWidth=Math.max(width,1e-6),safeHeight=Math.max(height,1e-6);let nextW=safeWidth,nextH=safeHeight;if(nextH/nextW>canvasAspect)nextW=nextH/canvasAspect;else nextH=nextW*canvasAspect;view={x:centerX-nextW/2,z:centerZ-nextH/2,w:nextW,h:nextH}}
function resetView(){const b=displayBounds(),w=b.maxX-b.minX,h=b.maxZ-b.minZ,pad=Math.max(w,h)*.045;fitViewToCanvas((b.minX+b.maxX)/2,(b.minZ+b.maxZ)/2,w+pad*2,h+pad*2);draw()}
function resize(){const r=canvas.getBoundingClientRect();canvas.width=Math.max(1,Math.floor(r.width*dpr));canvas.height=Math.max(1,Math.floor(r.height*dpr));if(view.w){fitViewToCanvas(view.x+view.w/2,view.z+view.h/2,view.w,view.h)}draw()}
function worldToScreen(x,z){return{x:(x-view.x)/view.w*canvas.width,y:(z-view.z)/view.h*canvas.height}}
function screenToWorld(x,y){return{x:view.x+x/canvas.width*view.w,z:view.z+y/canvas.height*view.h}}
function viewTransformCenter(){const b=stats.bounds;return{x:(b.minX+b.maxX)/2,z:(b.minZ+b.maxZ)/2}}
function rawToDisplay(x,z){const c=viewTransformCenter(),mirror=mirrorEl.value;let dx=x-c.x,dz=z-c.z;if(mirror==='x'||mirror==='xz')dx=-dx;if(mirror==='z'||mirror==='xz')dz=-dz;const angle=Number(rotationEl.value||0);if(angle===90)return{x:c.x-dz,z:c.z+dx};if(angle===180)return{x:c.x-dx,z:c.z-dz};if(angle===270)return{x:c.x+dz,z:c.z-dx};return{x:c.x+dx,z:c.z+dz}}
function displayToRaw(x,z){const c=viewTransformCenter(),angle=Number(rotationEl.value||0);let dx=x-c.x,dz=z-c.z;if(angle===90){const oldDx=dz;dz=-dx;dx=oldDx}else if(angle===180){dx=-dx;dz=-dz}else if(angle===270){const oldDx=-dz;dz=dx;dx=oldDx}const mirror=mirrorEl.value;if(mirror==='x'||mirror==='xz')dx=-dx;if(mirror==='z'||mirror==='xz')dz=-dz;return{x:c.x+dx,z:c.z+dz}}
function rawToScreen(x,z){const p=rawToDisplay(x,z);return worldToScreen(p.x,p.z)}
function currentRawQueryBbox(){const corners=[[view.x,view.z],[view.x+view.w,view.z],[view.x,view.z+view.h],[view.x+view.w,view.z+view.h]].map(([x,z])=>displayToRaw(x,z));const b=bboxForPoints(corners,p=>p.x,p=>p.z);return[b.minX,b.minZ,b.maxX,b.maxZ]}
function searchMatch(p){const q=searchEl.value.trim().toLowerCase();if(!q)return true;return valueText(p.sub_id).toLowerCase().includes(q)}
function pointVisible(p){return pointsEl.checked&&enabledGroups.has(groupValue(p))&&pointProvinceSelected(p)&&searchMatch(p)&&(!issuesEl.checked||(p.nav&&p.nav.issue))}
function currentViewBbox(){return[view.x,view.z,view.x+view.w,view.z+view.h]}
function connectivityStatusText(){if(!connectivityMode)return'';if(!connectivityStartPoly)return' | 连通性检查：点击任意 NavMesh 区域开始检查';return` | 连通性检查：起点 ${connectivityStartPoly}，连通 ${connectedPolyIds.size}/${polygons.length}`}
function draw(){if(!view.w)return;ctx.clearRect(0,0,canvas.width,canvas.height);ctx.fillStyle='#edf2f7';ctx.fillRect(0,0,canvas.width,canvas.height);drawGrid();let mesh=baseModeEl.value==='full'?drawPolygons():drawBoundary();drawConnectivityOverlay();if(!connectivityMode){drawStationDistanceLines();drawPoints();drawValidationHighlights()}statusEl.textContent=`${points.filter(pointVisible).length}/${points.length} points${connectivityMode?' (连通性检查中已临时隐藏)':''} | ${provinceSelectionText()} | ${mesh} ${baseModeEl.value} | X ${fmt(view.x)}..${fmt(view.x+view.w)} / Z ${fmt(view.z)}..${fmt(view.z+view.h)}${connectivityStatusText()}`}
function drawGrid(){const step=view.w>900?100:view.w>350?50:20;ctx.save();ctx.lineWidth=1*dpr;ctx.strokeStyle='#c9d2de';ctx.globalAlpha=.7;for(let x=Math.floor(view.x/step)*step;x<=view.x+view.w;x+=step){const p=worldToScreen(x,view.z);ctx.beginPath();ctx.moveTo(p.x,0);ctx.lineTo(p.x,canvas.height);ctx.stroke()}for(let z=Math.floor(view.z/step)*step;z<=view.z+view.h;z+=step){const p=worldToScreen(view.x,z);ctx.beginPath();ctx.moveTo(0,p.y);ctx.lineTo(canvas.width,p.y);ctx.stroke()}ctx.restore()}
function drawPolygons(){const visible=querySpatialIndex(polyIndex,currentRawQueryBbox());let n=0;for(const poly of visible){const active=provinceSelected(poly);ctx.beginPath();const f=rawToScreen(poly.verts[0][0],poly.verts[0][1]);ctx.moveTo(f.x,f.y);for(let i=1;i<poly.verts.length;i++){const p=rawToScreen(poly.verts[i][0],poly.verts[i][1]);ctx.lineTo(p.x,p.y)}ctx.closePath();ctx.fillStyle=active?colorFor(poly):'#cbd5e1';ctx.globalAlpha=active?.72:.16;ctx.fill();if(strokeEl.checked){ctx.globalAlpha=active?.42:.20;ctx.strokeStyle=active?'#1f2937':'#94a3b8';ctx.lineWidth=(active?Math.max(.55,Math.min(1.2,850/view.w)):.55)*dpr;ctx.stroke()}n++}ctx.globalAlpha=1;return n}
function drawConnectivityOverlay(){if(!connectivityMode||!connectivityStartPoly)return;const visible=querySpatialIndex(polyIndex,currentRawQueryBbox());ctx.save();for(const poly of visible){const state=connectivityColor(poly),isStart=String(poly.i)===connectivityStartPoly;ctx.beginPath();const f=rawToScreen(poly.verts[0][0],poly.verts[0][1]);ctx.moveTo(f.x,f.y);for(let i=1;i<poly.verts.length;i++){const p=rawToScreen(poly.verts[i][0],poly.verts[i][1]);ctx.lineTo(p.x,p.y)}ctx.closePath();ctx.fillStyle=state==='connected'?'#22c55e':'#ef4444';ctx.globalAlpha=state==='connected' ? .72 : .34;ctx.fill();ctx.globalAlpha=isStart ? .95 : .42;ctx.strokeStyle=state==='connected'?'#14532d':'#991b1b';ctx.lineWidth=(isStart?3.6:1.1)*dpr;ctx.stroke()}ctx.restore()}
function drawBoundary(){const visible=querySpatialIndex(boundaryIndex,currentRawQueryBbox()).filter(s=>baseModeEl.value!=='minimal'||(s.kind==='partition'&&s.crossProvinceBoundary)||(s.boundaryComponentArea||0)>=minimalBoundaryAreaThreshold);ctx.save();for(const s of visible){const active=provinceSelected(s);ctx.globalAlpha=active?.95:.24;ctx.lineWidth=(active?(baseModeEl.value==='minimal'?Math.max(1.4,Math.min(3.2,1800/view.w)):Math.max(1.2,Math.min(2.5,1400/view.w))):.75)*dpr;ctx.strokeStyle=active?colorFor(s):'#94a3b8';const a=rawToScreen(s.a[0],s.a[1]),b=rawToScreen(s.b[0],s.b[1]);ctx.beginPath();ctx.moveTo(a.x,a.y);ctx.lineTo(b.x,b.y);ctx.stroke()}ctx.restore();return visible.length}
function drawPoints(){if(connectivityMode||!pointsEl.checked)return;for(const pnt of points){if(!pointVisible(pnt))continue;const p=rawToScreen(pnt.x,pnt.z),hasIssue=pnt.nav&&pnt.nav.issue,color=pointColor(pnt),r=(hasIssue?8.5:6.5)*dpr;ctx.save();ctx.globalAlpha=.28;ctx.beginPath();ctx.arc(p.x,p.y,r+5*dpr,0,Math.PI*2);ctx.fillStyle=color;ctx.fill();ctx.globalAlpha=.95;ctx.beginPath();ctx.arc(p.x,p.y,r+2.2*dpr,0,Math.PI*2);ctx.fillStyle='#fff';ctx.fill();ctx.beginPath();ctx.arc(p.x,p.y,r,0,Math.PI*2);ctx.fillStyle=color;ctx.fill();ctx.globalAlpha=1;ctx.strokeStyle=hasIssue?'#111827':'#1f2937';ctx.lineWidth=(hasIssue?2.8:1.7)*dpr;ctx.stroke();if(selectedPoint&&selectedPoint.index===pnt.index){ctx.beginPath();ctx.arc(p.x,p.y,r+8*dpr,0,Math.PI*2);ctx.strokeStyle='#facc15';ctx.lineWidth=3*dpr;ctx.stroke()}ctx.restore()}}
function drawValidationHighlights(){if(!validationErrors.length)return;const seen=new Set();ctx.save();for(const error of validationErrors){const pnt=error.point;if(!pnt||seen.has(pnt.index))continue;seen.add(pnt.index);const p=rawToScreen(pnt.x,pnt.z),pulse=selectedPoint&&selectedPoint.index===pnt.index;ctx.beginPath();ctx.arc(p.x,p.y,(pulse?18:14)*dpr,0,Math.PI*2);ctx.strokeStyle='#ef4444';ctx.lineWidth=(pulse?4:3)*dpr;ctx.stroke();ctx.beginPath();ctx.arc(p.x,p.y,5*dpr,0,Math.PI*2);ctx.fillStyle='#ef4444';ctx.fill();ctx.strokeStyle='#fff';ctx.lineWidth=1.5*dpr;ctx.stroke()}ctx.restore()}
function drawStationDistanceLines(){if(!stationDistanceLines.length)return;const bbox=currentRawQueryBbox();ctx.save();for(const line of stationDistanceLines){const route=line.route&&line.route.length?line.route:[{x:line.a.x,z:line.a.z},{x:line.b.x,z:line.b.z}],rb=bboxForPoints(route,p=>p.x,p=>p.z);if(rb.maxX<bbox[0]||rb.minX>bbox[2]||rb.maxZ<bbox[1]||rb.minZ>bbox[3])continue;const top=stationDistanceTopIds.has(line.id);ctx.globalAlpha=top ? .95 : .18;ctx.strokeStyle=top?'#f97316':'#2563eb';ctx.lineWidth=(top?3.2:1.15)*dpr;ctx.beginPath();const first=rawToScreen(route[0].x,route[0].z);ctx.moveTo(first.x,first.y);for(let i=1;i<route.length;i++){const p=rawToScreen(route[i].x,route[i].z);ctx.lineTo(p.x,p.y)}ctx.stroke();if(top){const mid=route[Math.floor(route.length/2)],m=rawToScreen(mid.x,mid.z);ctx.globalAlpha=.96;ctx.fillStyle='rgba(255,255,255,.92)';ctx.strokeStyle='#f97316';ctx.lineWidth=1*dpr;const text=`${fmt(line.distance,0)}m / ${fmt(line.seconds,1)}s`,pad=4*dpr,w=ctx.measureText(text).width+pad*2,h=16*dpr;ctx.beginPath();ctx.rect(m.x-w/2,m.y-h/2,w,h);ctx.fill();ctx.stroke();ctx.fillStyle='#9a3412';ctx.font=`${12*dpr}px Segoe UI, Arial`;ctx.fillText(text,m.x-w/2+pad,m.y+4*dpr)}}ctx.restore();ctx.globalAlpha=1}
function pointInPoly(point,verts){let inside=false;for(let i=0,j=verts.length-1;i<verts.length;j=i++){const xi=verts[i][0],zi=verts[i][1],xj=verts[j][0],zj=verts[j][1];if((zi>point.z)!==(zj>point.z)&&point.x<(xj-xi)*(point.z-zi)/(zj-zi+1e-12)+xi)inside=!inside}return inside}
function findPoly(displayWorld){const world=displayToRaw(displayWorld.x,displayWorld.z);const candidates=querySpatialIndex(polyIndex,[world.x-1,world.z-1,world.x+1,world.z+1]);for(let i=candidates.length-1;i>=0;i--)if(pointInPoly(world,candidates[i].verts))return candidates[i];return null}
function findPoint(screen){let best=null,bestD=Infinity;for(const pnt of points){if(!pointVisible(pnt))continue;const p=rawToScreen(pnt.x,pnt.z),d=Math.hypot(p.x-screen.x,p.y-screen.y);if(d<bestD&&d<=18*dpr){best=pnt;bestD=d}}return best}
function selectPoint(p){selectedPoint=p||null;draw()}
function locatePointByIndex(v){const idx=Number(v);if(!Number.isInteger(idx))return;const p=points.find(x=>x.index===idx);if(!p)return;selectPoint(p);const display=rawToDisplay(p.x,p.z),targetW=Math.min(view.w,260);fitViewToCanvas(display.x,display.z,targetW,targetW);draw()}
function isTransferStation(p){return String(p.obj_id)==='8'}
function rotationText(p){return p.rot_x===null||p.rot_x===undefined||p.rot_w===null||p.rot_w===undefined?'-':`(${fmt(p.rot_x,4)}, ${fmt(p.rot_y??0,4)}, ${fmt(p.rot_z??0,4)}, ${fmt(p.rot_w,4)})`}
function forwardText(p){return p.forward_x===null||p.forward_x===undefined||p.forward_z===null||p.forward_z===undefined?'-':`(${fmt(p.forward_x,4)}, ${fmt(p.forward_y??0,4)}, ${fmt(p.forward_z,4)})`}
function stationForwardPoint(p,distance=5){const dx=Number(p.forward_x),dz=Number(p.forward_z);if(!Number.isFinite(dx)||!Number.isFinite(dz))return{issue:'MISSING_ROTATION'};const len=Math.hypot(dx,dz);if(len<1e-6)return{issue:'INVALID_FORWARD'};return{x:p.x+dx/len*distance,y:p.y,z:p.z+dz/len*distance,dirLen:len}}
function stationForwardNavIssue(p){const target=stationForwardPoint(p,5);if(target.issue)return{issue:target.issue,target:null,nav:null};const nav=classifyPoint(target);return nav.inside?null:{issue:'FORWARD_OUTSIDE_MESH',target,nav}}
function validationPointDetail(p){const poly=matchedPolygonForPoint(p);return `components[${p.index}] | ${rawTypeText(p)} | X/Z/Y ${fmt(p.x)}, ${fmt(p.z)}, ${fmt(p.y,3)} | Rotation ${rotationText(p)} | Forward ${forwardText(p)} | Poly ${p.nav?p.nav.poly:'-'} | ${poly?poly.areaKind||'未知区域':'未命中 NavMesh'}`}
function stationForwardDetail(p,result){const base=validationPointDetail(p);if(!result||!result.target)return `${base} | 前方5m: 无法计算，Rotation ${rotationText(p)} Forward ${forwardText(p)}`;return `${base} | 前方5m X/Z/Y ${fmt(result.target.x)}, ${fmt(result.target.z)}, ${fmt(result.target.y,3)} | ForwardPoly ${result.nav?result.nav.poly:'-'}`}
function setValidationProgress(done,total,label){const pct=total?Math.round(done/total*100):0;validationProgressBarEl.style.width=`${pct}%`;validationStatusEl.textContent=`${label} ${pct}% (${done}/${total})`}
function renderValidationResults(){if(!validationErrors.length){validationResultsEl.innerHTML='';validationStatusEl.textContent=validationProgressBarEl.style.width==='100%'?'检查完成，未发现错误点位。':'未开始检查。';draw();return}const grouped=countBy(validationErrors,e=>e.check);const summary=Array.from(grouped.entries()).map(([k,v])=>`${k} ${v}`).join('，');validationStatusEl.textContent=`检查完成，发现 ${validationErrors.length} 个错误点位：${summary}`;validationResultsEl.innerHTML=validationErrors.map((e,i)=>`<button class="error-item" data-error-index="${i}"><b>${escapeHtml(e.check)}：components[${e.point.index}]</b><span>${escapeHtml(e.message)}</span><span>${escapeHtml(e.detail)}</span></button>`).join('');validationResultsEl.querySelectorAll('.error-item').forEach(btn=>btn.addEventListener('click',()=>{const error=validationErrors[Number(btn.dataset.errorIndex)];if(error)locatePointByIndex(error.point.index)}));draw()}
function clearValidationResults(){validationErrors=[];validationProgressBarEl.style.width='0%';validationResultsEl.innerHTML='';validationStatusEl.textContent='未开始检查。';draw()}
async function runValidationChecks(){
if(validationRunning)return;
const checkNavmesh=document.getElementById('check-navmesh-overlap').checked;
const checkStationValidity=document.getElementById('check-station-validity').checked;
if(!checkNavmesh&&!checkStationValidity){validationStatusEl.textContent='请至少选择一个检查项。';return}
validationRunning=true;validationErrors=[];validationResultsEl.innerHTML='';validationProgressBarEl.style.width='0%';
const total=points.length*(Number(checkNavmesh)+Number(checkStationValidity));let done=0;
try{
  if(checkNavmesh){
    for(let i=0;i<points.length;i++){
      const p=points[i];p.nav=classifyPoint(p);
      if(!p.nav.inside)validationErrors.push({check:'点位/NavMesh',point:p,message:'点位未落在当前 NavMesh 上。',detail:validationPointDetail(p)});
      done++;if(i%25===0){setValidationProgress(done,total,'当前检查项：点位是否落在 NavMesh 上');await new Promise(resolve=>setTimeout(resolve,0))}
    }
  }
  if(checkStationValidity){
    for(let i=0;i<points.length;i++){
      const p=points[i];if(!p.nav)p.nav=classifyPoint(p);
      if(isTransferStation(p)){
        const poly=matchedPolygonForPoint(p);
        if(poly&&poly.areaKind==='PVP区域')validationErrors.push({check:'传送车站合法性',point:p,message:'传送车站位于 PVP 区域。',detail:validationPointDetail(p)});
        const result=stationForwardNavIssue(p);
        if(result){
          const message=result.issue==='MISSING_ROTATION'?'传送车站缺少 rotation 四元数。':result.issue==='INVALID_FORWARD'?'传送车站 rotation 换算出的 X/Z 朝向长度为 0。':'传送车站朝向前方 5m 不在 NavMesh 内。';
          validationErrors.push({check:'传送车站合法性',point:p,message,detail:stationForwardDetail(p,result)});
        }
      }
      done++;if(i%25===0){setValidationProgress(done,total,'当前检查项：传送车站合法性检查');await new Promise(resolve=>setTimeout(resolve,0))}
    }
  }
  setValidationProgress(total,total,'检查完成');renderValidationResults();
}finally{validationRunning=false}
}
function setStationDistanceProgress(done,total,label){const pct=total?Math.round(done/total*100):0;stationDistanceProgressBarEl.style.width=`${pct}%`;stationDistanceStatusEl.textContent=`${label} ${pct}% (${done}/${total})`}
function clearStationDistanceResults(){stationDistanceLines=[];stationDistanceTopIds=new Set();stationDistanceProgressBarEl.style.width='0%';stationDistanceResultsEl.innerHTML='';stationDistanceStatusEl.textContent='未开始检查。';draw()}
function buildNavGraph(){if(navGraph)return navGraph;stationDistanceStatusEl.textContent='基于完整 NavMesh polygon 构建寻路图...';const nodeCount=polygons.length,polyToNode=new Map(),centers=new Array(nodeCount),adj=Array.from({length:nodeCount},()=>[]),edges=new Map();for(let i=0;i<polygons.length;i++){const poly=polygons[i],xs=poly.verts.map(v=>v[0]),zs=poly.verts.map(v=>v[1]);polyToNode.set(String(poly.i),i);centers[i]={x:xs.reduce((a,b)=>a+b,0)/xs.length,z:zs.reduce((a,b)=>a+b,0)/zs.length};const vs=poly.verts;for(let n=0;n<vs.length;n++){const a=vs[n],b=vs[(n+1)%vs.length],key=[`${a[0].toFixed(3)},${a[1].toFixed(3)}`,`${b[0].toFixed(3)},${b[1].toFixed(3)}`].sort().join('|');if(!edges.has(key))edges.set(key,[]);edges.get(key).push(i)}}for(const members of edges.values()){if(members.length<2)continue;for(let a=0;a<members.length;a++)for(let b=a+1;b<members.length;b++){const i=members[a],j=members[b],ci=centers[i],cj=centers[j],w=Math.hypot(ci.x-cj.x,ci.z-cj.z);adj[i].push([j,w]);adj[j].push([i,w])}}navGraph={adj,centers,polyToNode};return navGraph}
class MinHeap{constructor(){this.a=[]}push(item){this.a.push(item);let i=this.a.length-1;while(i>0){const p=(i-1)>>1;if(this.a[p][0]<=item[0])break;this.a[i]=this.a[p];i=p}this.a[i]=item}pop(){if(!this.a.length)return null;const root=this.a[0],last=this.a.pop();if(this.a.length&&last){let i=0;while(true){let l=i*2+1,r=l+1;if(l>=this.a.length)break;let c=r<this.a.length&&this.a[r][0]<this.a[l][0]?r:l;if(this.a[c][0]>=last[0])break;this.a[i]=this.a[c];i=c}this.a[i]=last}return root}get length(){return this.a.length}}
function stationLabel(p){return `${p.obj_display_name||rawTypeText(p)} components[${p.index}]`}
function routeStationToNode(station,node,prev,graph){const nodes=[];let cur=node,guard=0;while(cur!==-1&&guard++<graph.centers.length+2){nodes.push(cur);if(cur===station.node)break;cur=prev[cur]}nodes.reverse();const route=[{x:station.x,z:station.z}];for(const n of nodes){const c=graph.centers[n];route.push({x:c.x,z:c.z})}return route}
function routeBetweenStations(a,b,leftNode,rightNode,prev,graph){if(leftNode===rightNode)return[{x:a.x,z:a.z},{x:graph.centers[leftNode].x,z:graph.centers[leftNode].z},{x:b.x,z:b.z}];const left=routeStationToNode(a,leftNode,prev,graph),right=routeStationToNode(b,rightNode,prev,graph).reverse();return left.concat(right)}
function renderStationDistanceResults(){if(!stationDistanceLines.length){stationDistanceResultsEl.innerHTML='';stationDistanceStatusEl.textContent='没有可显示的车站最近邻结果。';draw();return}const top=stationDistanceLines.filter(l=>stationDistanceTopIds.has(l.id)).sort((a,b)=>b.distance-a.distance);stationDistanceStatusEl.textContent=`检查完成：基于完整 NavMesh polygon 计算出 ${stationDistanceLines.length} 条合并后的最近邻路线，已高亮其中距离最大的 ${top.length} 条。`;stationDistanceResultsEl.innerHTML=top.map((line,i)=>`<button class="distance-item" data-line-id="${line.id}"><b>#${i+1} ${fmt(line.distance,1)}m / ${fmt(line.seconds,1)}s</b><span>车站 A：${escapeHtml(stationLabel(line.a))}</span><span>车站 B：${escapeHtml(stationLabel(line.b))}</span></button>`).join('');stationDistanceResultsEl.querySelectorAll('.distance-item').forEach(btn=>btn.addEventListener('click',()=>{const line=stationDistanceLines.find(x=>String(x.id)===btn.dataset.lineId);if(!line)return;const route=line.route&&line.route.length?line.route:[{x:line.a.x,z:line.a.z},{x:line.b.x,z:line.b.z}],display=route.map(p=>rawToDisplay(p.x,p.z)),b=bboxForPoints(display,p=>p.x,p=>p.z),w=Math.max(b.maxX-b.minX,120),h=Math.max(b.maxZ-b.minZ,120),pad=Math.max(w,h)*.35;fitViewToCanvas((b.minX+b.maxX)/2,(b.minZ+b.maxZ)/2,w+pad*2,h+pad*2);selectedPoint=line.a;draw()}));draw()}
async function runStationDistanceCheck(){if(stationDistanceRunning)return;stationDistanceRunning=true;stationDistanceLines=[];stationDistanceTopIds=new Set();stationDistanceResultsEl.innerHTML='';stationDistanceProgressBarEl.style.width='0%';try{const graph=buildNavGraph(),stations=[],stationsByNode=new Map();for(const p of points){if(!isTransferStation(p))continue;if(!p.nav)p.nav=classifyPoint(p);if(!p.nav.inside)continue;const node=graph.polyToNode.get(String(p.nav.poly));if(node===undefined)continue;const station={...p,node,stationIndex:stations.length};stations.push(station);const list=stationsByNode.get(node)||[];list.push(station);stationsByNode.set(node,list)}if(stations.length<2){stationDistanceStatusEl.textContent=`可寻路传送车站不足 2 个：${stations.length}`;return}setStationDistanceProgress(0,3,'构建车站最近邻候选');const bestByStation=new Map(),updateBest=(a,b,distance,route)=>{if(!a||!b||a.stationIndex===b.stationIndex||!Number.isFinite(distance))return;const safeRoute=route&&route.length?route:[{x:a.x,z:a.z},{x:b.x,z:b.z}],lineAB={id:`${a.index}-${b.index}`,a,b,distance,seconds:distance/5.8,route:safeRoute},lineBA={id:`${b.index}-${a.index}`,a:b,b:a,distance,seconds:distance/5.8,route:safeRoute.slice().reverse()};const oldA=bestByStation.get(a.stationIndex),oldB=bestByStation.get(b.stationIndex);if(!oldA||distance<oldA.distance)bestByStation.set(a.stationIndex,lineAB);if(!oldB||distance<oldB.distance)bestByStation.set(b.stationIndex,lineBA)};for(const sameNodeStations of stationsByNode.values()){if(sameNodeStations.length<2)continue;for(let i=0;i<sameNodeStations.length;i++)for(let j=i+1;j<sameNodeStations.length;j++){const a=sameNodeStations[i],b=sameNodeStations[j],c=graph.centers[a.node];updateBest(a,b,Math.hypot(a.x-b.x,a.z-b.z),[{x:a.x,z:a.z},{x:c.x,z:c.z},{x:b.x,z:b.z}])}}setStationDistanceProgress(1,3,'多源 Dijkstra 扩散');const dist=new Float64Array(graph.adj.length),owner=new Int32Array(graph.adj.length),prev=new Int32Array(graph.adj.length),heap=new MinHeap();dist.fill(Infinity);owner.fill(-1);prev.fill(-1);for(const station of stations){const c=graph.centers[station.node],startDistance=Math.hypot(station.x-c.x,station.z-c.z);if(startDistance<dist[station.node]){dist[station.node]=startDistance;owner[station.node]=station.stationIndex;prev[station.node]=-1;heap.push([startDistance,station.node])}}let settled=0;while(heap.length){const cur=heap.pop(),d=cur[0],node=cur[1];if(d!==dist[node])continue;settled++;const nodeOwner=owner[node];for(const [next,w] of graph.adj[node]){const nextOwner=owner[next];if(nextOwner!==-1&&nextOwner!==nodeOwner){const a=stations[nodeOwner],b=stations[nextOwner],route=routeBetweenStations(a,b,node,next,prev,graph);updateBest(a,b,d+w+dist[next],route)}const nd=d+w;if(nd<dist[next]){dist[next]=nd;owner[next]=nodeOwner;prev[next]=node;heap.push([nd,next])}}if(settled%5000===0){setStationDistanceProgress(Math.min(settled,graph.adj.length),graph.adj.length,'多源 Dijkstra 扩散');await new Promise(resolve=>setTimeout(resolve,0))}}setStationDistanceProgress(2,3,'整理最近邻路线');const merged=new Map();for(const line of bestByStation.values()){const key=[line.a.index,line.b.index].sort((a,b)=>a-b).join('-');const old=merged.get(key);if(!old||line.distance<old.distance)merged.set(key,line)}stationDistanceLines=Array.from(merged.values()).sort((a,b)=>b.distance-a.distance);setStationDistanceProgress(3,3,'车站最近邻检查完成');const top=stationDistanceLines.slice().sort((a,b)=>b.distance-a.distance).slice(0,10);stationDistanceTopIds=new Set(top.map(l=>l.id));renderStationDistanceResults()}finally{stationDistanceRunning=false}}
function renderFilters(){const q=filterQuery.trim().toLowerCase(),visible=filterEntries.filter(e=>!q||e.key.toLowerCase().includes(q)||e.label.toLowerCase().includes(q)||e.category.toLowerCase().includes(q));const categoryTotals=new Map();for(const e of filterEntries)categoryTotals.set(e.category,(categoryTotals.get(e.category)||0)+e.count);const grouped=new Map();for(const e of visible){if(!grouped.has(e.category))grouped.set(e.category,[]);grouped.get(e.category).push(e)}const categories=Array.from(grouped.keys()).sort((a,b)=>(categoryTotals.get(b)||0)-(categoryTotals.get(a)||0)||a.localeCompare(b));filtersEl.innerHTML=categories.map(category=>{const entries=grouped.get(category).sort((a,b)=>b.count-a.count||a.label.localeCompare(b.label)),collapsed=collapsedFilterCategories.has(category);return `<div class="filter-category"><button class="filter-category-title" type="button" data-category="${escapeHtml(category)}"><span>${collapsed?'▸':'▾'}</span><span>${escapeHtml(category)}</span><b class="filter-category-count">${categoryTotals.get(category)||0}</b></button><div class="filter-category-body" style="${collapsed?'display:none':''}">${entries.map(e=>`<label class="filter-row"><input type="checkbox" data-key="${escapeHtml(e.key)}" ${enabledGroups.has(e.key)?'checked':''}><span><span class="swatch" style="background:${hashColor(e.key)}"></span> ${escapeHtml(e.label)}</span><b>${e.count}</b></label>`).join('')}</div></div>`}).join('')||'<p class="province-count">没有匹配点位分类</p>';filtersEl.querySelectorAll('.filter-category-title').forEach(btn=>btn.addEventListener('click',()=>{const category=btn.dataset.category;collapsedFilterCategories.has(category)?collapsedFilterCategories.delete(category):collapsedFilterCategories.add(category);renderFilters()}));filtersEl.querySelectorAll('input[data-key]').forEach(input=>input.addEventListener('change',()=>{input.checked?enabledGroups.add(input.dataset.key):enabledGroups.delete(input.dataset.key);fillStats();draw()}))}
function rebuildRawRows(){const counts=new Map(),labels=new Map();for(const p of points){const k=groupValue(p);counts.set(k,(counts.get(k)||0)+1);if(!labels.has(k)||labels.get(k)===k)labels.set(k,pointDisplayText(p))}return Array.from(counts.entries()).map(([key,count])=>({key,count,label:labels.get(key)||key})).sort((a,b)=>b.count-a.count||a.label.localeCompare(b.label))}
function rebuildFilters(reset=true){const counts=new Map(),categoryByKey=new Map(),labelByKey=new Map();for(const p of points){const k=groupValue(p),category=pointFilterCategory(p),label=pointDisplayText(p);counts.set(k,(counts.get(k)||0)+1);if(!categoryByKey.has(k))categoryByKey.set(k,category);if(!labelByKey.has(k)||labelByKey.get(k)===k)labelByKey.set(k,label)}filterEntries=Array.from(counts.entries()).map(([key,count])=>({key,count,category:categoryByKey.get(key)||'Other',label:labelByKey.get(key)||key})).sort((a,b)=>b.count-a.count||a.label.localeCompare(b.label));if(reset)enabledGroups=new Set(filterEntries.map(x=>x.key));renderFilters();draw()}
function escapeHtml(v){return String(v).replace(/[&<>"']/g,ch=>({'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&#39;'}[ch]))}
function countBy(items,getter){const m=new Map();for(const item of items){const k=valueText(getter(item));m.set(k,(m.get(k)||0)+1)}return m}
function updateRawTypeToggle(){const collapsed=rawTypeTableEl.classList.contains('collapsed');rawTypeToggleEl.textContent=collapsed?'展开全部':'收起';rawTypeToggleEl.style.display=rawTypeTableEl.rows.length>5?'inline-flex':'none'}
function fillStats(){const b=stats.bounds,summary=analysis.summary||{},outside=points.filter(p=>p.nav&&!p.nav.inside).length,isolated=points.filter(p=>p.nav&&p.nav.issue==='ISOLATED_COMPONENT').length,visibleCount=points.filter(pointVisible).length;document.getElementById('metrics').innerHTML=[['MapWidth / Height',`${fmt(stats.mapWidth)} / ${fmt(stats.mapHeight)}`],['Polygons',polygons.length],['Components',`${stats.componentCount||'-'}, main ${stats.mainComponentSize||'-'}`],['Boundary segments',boundarySegments.length],['X range',`${fmt(b.minX)}..${fmt(b.maxX)}`],['Z range',`${fmt(b.minZ)}..${fmt(b.maxZ)}`],['LevelPoints',`${visibleCount} visible / ${summary.raw_component_count||'-'} raw`],['Province filter',`${selectedProvinces.size}/${allProvinceIds.length||0}`],['Skipped no position',summary.skipped_component_count||0],['Outside / isolated',`${outside} / ${isolated}`]].map(([k,v])=>`<div class="metric"><span>${k}</span><b>${v}</b></div>`).join('');const labelRows=Array.from(countBy(points,p=>p.label).entries()).sort((a,b)=>b[1]-a[1]).slice(0,20),classRows=Array.from(countBy(points,p=>p.class_name).entries()).sort((a,b)=>b[1]-a[1]).slice(0,20),rawRows=filterEntries.length?filterEntries:rebuildRawRows();rawTypeTableEl.innerHTML=rawRows.map(e=>`<tr><td><span class="swatch" style="display:inline-block;background:${hashColor(e.key)}"></span> ${escapeHtml(e.label)}</td><td>${e.count}</td></tr>`).join('');updateRawTypeToggle();document.getElementById('label-table').innerHTML=labelRows.map(([k,v])=>`<tr><td>${escapeHtml(k)}</td><td>${v}</td></tr>`).join('');document.getElementById('class-table').innerHTML=classRows.map(([k,v])=>`<tr><td>${escapeHtml(k)}</td><td>${v}</td></tr>`).join('');navmeshPathEl.value=stats.source||'';levelpointsPathEl.value=stats.levelpointsSource||summary.levelpoints||''}
function setupElectronControls(){if(!window.levelpointsElectron)return;document.getElementById('choose-navmesh').textContent='设置根目录';document.getElementById('choose-levelpoints').style.display='none';sourceStatusEl.textContent='Electron 桌面版：只需要设置一次根目录，NavMesh、LevelPoints、InteractiveObj 和 area_config 会按固定相对路径自动查找。'}
function parseNavmeshJson(nav,name){setSourceProgress(50,'解析 NavMesh：生成 polygon、连通分量和边界索引');const next=(nav.NavMeshPolygons||[]).map((poly,i)=>polygonRecord(poly,i)).filter(p=>p.verts.length>=3);applyAreaConfig(next);addComponents(next);polygons=next;boundarySegments=addBoundaryComponentAreas(makeBoundarySegments(next));const componentIds=new Set(),componentSizes=minMaxValues(next,p=>p.componentSize||0);for(const p of next)componentIds.add(p.component);stats={...stats,source:name,polygonCount:next.length,bounds:boundsFor(next),componentCount:componentIds.size,mainComponentSize:componentSizes.max,boundarySegmentCount:boundarySegments.length,minimalBoundaryAreaThreshold};navGraph=null;connectivityGraph=null;connectivityMode=false;connectivityStartPoly=null;connectedPolyIds=new Set();document.body.classList.remove('connectivity-mode');connectivityToggleEl.textContent='连通性检查';connectivityToggleEl.style.background='';rebuildIndexes();rebuildProvinceFilter(true);setSourceProgress(58,'解析 NavMesh：重新检查现有点位归属');for(const p of points)p.nav=classifyPoint(p);clearValidationResults();clearStationDistanceResults();fillStats();rebuildFilters(false);resetView();sourceStatusEl.textContent=`已加载 NavMesh：${name}`}
function polygonRecord(poly,i){const rawVerts=poly.Vertexs||[],verts=rawVerts.map(v=>[Number(v.x||0),Number(v.z||0)]);let ySum=0;for(const v of rawVerts)ySum+=Number(v.y||0);const b=bboxForPoints(verts,v=>v[0],v=>v[1]);return{i:poly.PolygonIndex??i,area:poly.AreaType,group:poly.GroupIndex,province:poly.ProvinceID,land:poly.LandID,tile:poly.TileIndex,y:rawVerts.length?ySum/rawVerts.length:0,verts,bbox:verts.length?[b.minX,b.minZ,b.maxX,b.maxZ]:[0,0,0,0]}}
function applyAreaConfig(polys){for(const poly of polys){const info=areaConfig[String(poly.province)]||areaConfig[poly.province]||{};poly.areaKind=info.areaKind||'未知区域';poly.areaName=info.areaName||'';poly.gveType=info.gveType;poly.pvpType=info.pvpType}}
function addComponents(polys){const edges=new Map();polys.forEach((poly,i)=>{const vs=poly.verts;for(let n=0;n<vs.length;n++){const a=vs[n],b=vs[(n+1)%vs.length],key=[`${a[0].toFixed(3)},${a[1].toFixed(3)}`,`${b[0].toFixed(3)},${b[1].toFixed(3)}`].sort().join('|');if(!edges.has(key))edges.set(key,[]);edges.get(key).push(i)}});const graph=polys.map(()=>[]);for(const members of edges.values())if(members.length>1)for(const i of members)for(const j of members)if(j!==i)graph[i].push(j);const comp=Array(polys.length).fill(-1),sizes=[];for(let s=0;s<polys.length;s++){if(comp[s]!==-1)continue;const cid=sizes.length,q=[s];comp[s]=cid;let c=0;while(q.length){const cur=q.pop();c++;for(const nxt of graph[cur])if(comp[nxt]===-1){comp[nxt]=cid;q.push(nxt)}}sizes.push(c)}let main=0,biggest=-1;for(let i=0;i<sizes.length;i++)if(sizes[i]>biggest){biggest=sizes[i];main=i}polys.forEach((p,i)=>{p.component=comp[i];p.componentSize=sizes[comp[i]];p.mainComponent=comp[i]===main})}
function regionKey(poly){return [poly.province,poly.group,poly.component,poly.area].join('|')}
function makeBoundarySegments(polys){const edges=new Map();for(const poly of polys){const vs=poly.verts;for(let i=0;i<vs.length;i++){const a=vs[i],b=vs[(i+1)%vs.length],key=edgeKey(a,b);if(!edges.has(key))edges.set(key,[]);edges.get(key).push({a,b,poly})}}const out=[];for(const members of edges.values()){const sig=new Set(members.map(m=>regionKey(m.poly)));if(members.length!==1&&sig.size<=1)continue;const m=members[0],provinces=new Set(members.map(x=>valueText(x.poly.province)));out.push({a:m.a,b:m.b,bbox:[Math.min(m.a[0],m.b[0]),Math.min(m.a[1],m.b[1]),Math.max(m.a[0],m.b[0]),Math.max(m.a[1],m.b[1])],kind:members.length===1?'outer':'partition',...m.poly,vertexCount:m.poly.verts.length,regionKeys:Array.from(sig).sort(),crossProvinceBoundary:members.length>1&&provinces.size>1,adjacentProvinces:Array.from(provinces).sort()})}return out}
function boundaryMembersArea(segments,members){let minX=Infinity,maxX=-Infinity,minZ=Infinity,maxZ=-Infinity;for(const idx of members){const s=segments[idx];for(const p of [s.a,s.b]){const x=p[0],z=p[1];if(x<minX)minX=x;if(x>maxX)maxX=x;if(z<minZ)minZ=z;if(z>maxZ)maxZ=z}}return minX===Infinity?0:(maxX-minX)*(maxZ-minZ)}
function addBoundaryComponentAreas(segments){
const sideEndpointToSegments=new Map(),outerEndpointToSegments=new Map();
segments.forEach((segment,i)=>{
  if(segment.kind==='outer'){
    for(const point of [segment.a,segment.b]){
      const key=`${point[0].toFixed(3)},${point[1].toFixed(3)}`;
      if(!outerEndpointToSegments.has(key))outerEndpointToSegments.set(key,[]);
      outerEndpointToSegments.get(key).push(i);
    }
    return;
  }
  for(const region of segment.regionKeys||[]){
    for(const point of [segment.a,segment.b]){
      const key=`${region}|${point[0].toFixed(3)},${point[1].toFixed(3)}`;
      if(!sideEndpointToSegments.has(key))sideEndpointToSegments.set(key,[]);
      sideEndpointToSegments.get(key).push(i);
    }
  }
});
const visitedOuter=new Set();
for(let start=0;start<segments.length;start++){
  if(segments[start].kind!=='outer'||visitedOuter.has(start))continue;
  const stack=[start],members=[];
  visitedOuter.add(start);
  while(stack.length){
    const cur=stack.pop();
    members.push(cur);
    for(const point of [segments[cur].a,segments[cur].b]){
      const key=`${point[0].toFixed(3)},${point[1].toFixed(3)}`;
      for(const next of outerEndpointToSegments.get(key)||[]){
        if(!visitedOuter.has(next)){visitedOuter.add(next);stack.push(next)}
      }
    }
  }
  const area=boundaryMembersArea(segments,members);
  for(const idx of members){segments[idx].boundaryComponentArea=area;segments[idx].boundaryComponentSize=members.length}
}
const visited=new Set();
for(let start=0;start<segments.length;start++){
  if(segments[start].kind==='outer')continue;
  for(const region of segments[start].regionKeys||[]){
    const startState=`${region}|${start}`;
    if(visited.has(startState))continue;
    const stack=[start],members=[];
    visited.add(startState);
    while(stack.length){
      const cur=stack.pop();
      members.push(cur);
      for(const point of [segments[cur].a,segments[cur].b]){
        const key=`${region}|${point[0].toFixed(3)},${point[1].toFixed(3)}`;
        for(const next of sideEndpointToSegments.get(key)||[]){
          const state=`${region}|${next}`;
          if(!visited.has(state)){visited.add(state);stack.push(next)}
        }
      }
    }
    const area=boundaryMembersArea(segments,members);
    for(const idx of members){
      const current=segments[idx].boundaryComponentArea;
      segments[idx].boundaryComponentArea=current===undefined?area:Math.min(current,area);
      segments[idx].boundaryComponentSize=Math.max(segments[idx].boundaryComponentSize||0,members.length);
    }
  }
}
return segments}
function boundsFor(polys){let minX=Infinity,maxX=-Infinity,minZ=Infinity,maxZ=-Infinity,minY=Infinity,maxY=-Infinity;for(const poly of polys){if(Number.isFinite(poly.y)){if(poly.y<minY)minY=poly.y;if(poly.y>maxY)maxY=poly.y}for(const v of poly.verts){const x=v[0],z=v[1];if(Number.isFinite(x)){if(x<minX)minX=x;if(x>maxX)maxX=x}if(Number.isFinite(z)){if(z<minZ)minZ=z;if(z>maxZ)maxZ=z}}}return{minX:minX===Infinity?0:minX,maxX:maxX===-Infinity?0:maxX,minZ:minZ===Infinity?0:minZ,maxZ:maxZ===-Infinity?0:maxZ,minY:minY===Infinity?0:minY,maxY:maxY===-Infinity?0:maxY}}
function nestedMb(c){return(((c||{}).serializedData||{}).MonoBehaviour||{})}
function labelFor(p){return p.class_name==='NpcPoint'?'NPC':p.class_name==='TeamPoint'?'Team':p.class_name==='PetMonsterPoint'?'PetMonster':p.class_name==='InteractiveObjPoint'?'InteractiveObj':p.class_name||'Unknown'}
function rotationFromComponent(c){const r=c.rotation;if(!r||typeof r!=='object')return null;const x=Number(r.x??0),y=Number(r.y??0),z=Number(r.z??0),w=Number(r.w??1);return Number.isFinite(x)&&Number.isFinite(y)&&Number.isFinite(z)&&Number.isFinite(w)?{x,y,z,w}:null}
function forwardFromRotation(r){if(!r)return null;return{x:2*(r.x*r.z+r.w*r.y),y:2*(r.y*r.z-r.w*r.x),z:1-2*(r.x*r.x+r.y*r.y)}}
function objDisplayNameFromPrevious(objId){if(objId===null||objId===undefined)return null;const match=points.find(p=>p.obj_id===objId&&p.obj_display_name);return match?match.obj_display_name:null}
function labelFromPreviousObj(objId){if(objId===null||objId===undefined)return null;const match=points.find(p=>p.obj_id===objId&&p.label);return match?match.label:null}
function normalizePoint(c,i){if(!c.position)return null;const mb=nestedMb(c),prev=points.find(p=>p.index===i)||{},rot=rotationFromComponent(c),forward=forwardFromRotation(rot),objId=mb.ObjID??null,objName=prev.obj_id===objId?prev.obj_display_name:(objDisplayNameFromPrevious(objId)||interactiveObjName(objId));const p={index:i,class_name:c.className,sub_id:c.subID??mb.SubID,parent_sub_id:c.parentSubID,belong_to_layer_id:c.belongToLayerID??mb.BelongToLayerID,prefab_path:c.prefabPath||'',obj_id:objId,obj_type:mb.ObjType??null,npc_id:mb.NpcID??null,pet_monster_id:mb.PetMonsterID??null,team_id:mb.TeamID??null,area_id:mb.AreaID??null,x:Number(c.position.x||0),y:Number(c.position.y||0),z:Number(c.position.z||0),rot_x:rot?rot.x:null,rot_y:rot?rot.y:null,rot_z:rot?rot.z:null,rot_w:rot?rot.w:null,forward_x:forward?forward.x:null,forward_y:forward?forward.y:null,forward_z:forward?forward.z:null,obj_display_name:objName};p.label=prev.obj_id===objId&&prev.label?prev.label:(labelFromPreviousObj(objId)||labelFor(p));p.nav=classifyPoint(p);return p}
function classifyPoint(p){const candidates=querySpatialIndex(polyIndex,[p.x-.001,p.z-.001,p.x+.001,p.z+.001]);for(const poly of candidates)if(pointInPoly(p,poly.verts))return{inside:true,poly:poly.i,component:poly.component,main:!!poly.mainComponent,issue:poly.mainComponent?null:'ISOLATED_COMPONENT'};return{inside:false,poly:null,component:null,main:false,issue:'OUTSIDE_MESH'}}
function parseLevelpointsJson(json,name){setSourceProgress(75,'分析 LevelPoints：解析 components 并检查 NavMesh 归属');const components=Array.isArray(json.components)?json.components:[],next=[];for(let i=0;i<components.length;i++){const p=normalizePoint(components[i],i);if(p)next.push(p);if(i&&i%100===0)setSourceProgress(75+Math.min(18,i/Math.max(components.length,1)*18),`分析 LevelPoints：${i}/${components.length}`)}points=next;analysis.summary={...(analysis.summary||{}),levelpoints:name,raw_component_count:components.length,point_count:points.length,skipped_component_count:components.length-points.length};stats.levelpointsSource=name;selectedPoint=null;selectPoint(null);clearValidationResults();clearStationDistanceResults();rebuildFilters(true);fillStats();draw();setSourceProgress(96,'分析 LevelPoints：刷新筛选、统计和地图显示');sourceStatusEl.textContent=`已加载 LevelPoints：${name}，${points.length}/${components.length} 个有 position 的点位。`}
function setSourceProgress(percent,label){sourceProgressBarEl.style.width=`${Math.max(0,Math.min(100,percent))}%`;sourceProgressStatusEl.textContent=label}
function fileStamp(file){return `${file.name} | ${Math.round((file.size||0)/1024)}KB | ${file.lastModified?new Date(file.lastModified).toLocaleString():'未知修改时间'}`}
function readJsonFileAsync(file,label,start=0,end=100){return new Promise((resolve,reject)=>{const r=new FileReader();r.onprogress=e=>{if(e.lengthComputable)setSourceProgress(start+(end-start)*e.loaded/e.total,`${label}：读取中 ${Math.round(e.loaded/Math.max(e.total,1)*100)}%`)};r.onload=()=>{try{setSourceProgress(end,`${label}：解析中`);resolve({json:JSON.parse(String(r.result||'')),name:file.name,file})}catch(e){reject(e)}};r.onerror=()=>reject(r.error||new Error(`${label} 读取失败`));r.readAsText(file,'utf-8')})}
async function ensureHandlePermission(handle,label){if(!handle||!handle.queryPermission)return true;const opts={mode:'read'};let state=await handle.queryPermission(opts);if(state==='granted')return true;setSourceProgress(5,`${label}：请求读取权限`);state=await handle.requestPermission(opts);if(state!=='granted')throw new Error(`${label} 没有读取权限，请重新选择文件授权。`);return true}
async function readJsonFromHandle(kind,label,start=0,end=100){const handle=kind==='navmesh'?lastNavmeshHandle:lastLevelpointsHandle;if(!handle)return null;await ensureHandlePermission(handle,label);setSourceProgress(start,`${label}：通过文件句柄读取磁盘最新内容`);const latest=await handle.getFile();if(kind==='navmesh')lastNavmeshFile=latest;else lastLevelpointsFile=latest;const loaded=await readJsonFileAsync(latest,label,start,end);loaded.name=latest.name;loaded.stamp=fileStamp(latest);return loaded}
async function promptReselectSnapshot(kind,label){sourceStatusEl.textContent=`${label} 当前只有浏览器 File 快照，无法读取磁盘最新内容。请重新选择文件以获取最新版本。`;setSourceProgress(0,`${label}：需要重新选择文件`);await chooseSourceFile(kind);return null}
async function refreshSource(){try{if(window.levelpointsElectron){setSourceProgress(0,'Electron：按当前路径读取磁盘最新文件');const loaded=await window.levelpointsElectron.reloadSources();if(loaded.navmesh){setSourceProgress(30,'Electron：解析最新 NavMesh');parseNavmeshJson(loaded.navmesh.json,loaded.navmesh.name||loaded.navmesh.path)}if(loaded.levelpoints){setSourceProgress(72,'Electron：解析最新 LevelPoints');parseLevelpointsJson(loaded.levelpoints.json,loaded.levelpoints.name||loaded.levelpoints.path)}setSourceProgress(100,'Electron：重新加载完成');sourceStatusEl.textContent='Electron 桌面版已按当前路径重新读取磁盘最新文件。';return}if(!(lastNavmeshHandle||lastLevelpointsHandle||lastNavmeshFile||lastLevelpointsFile)){sourceStatusEl.textContent='尚未授权源文件。请在弹出的选择器中选择最新 LevelPoints.json，选择后会立即重跑点位分析。';setSourceProgress(0,'等待选择 LevelPoints.json');await chooseSourceFile('levelpoints');return}setSourceProgress(0,'开始重新加载已授权源文件');if(lastNavmeshHandle&&lastLevelpointsHandle){const nav=await readJsonFromHandle('navmesh','NavMesh',0,45);parseNavmeshJson(nav.json,nav.name);const level=await readJsonFromHandle('levelpoints','LevelPoints',45,95);parseLevelpointsJson(level.json,level.name);setSourceProgress(100,'重新加载完成：NavMesh + LevelPoints');sourceStatusEl.textContent=`已读取磁盘最新文件：NavMesh ${nav.stamp}；LevelPoints ${level.stamp}`;return}if(lastNavmeshHandle){const nav=await readJsonFromHandle('navmesh','NavMesh',0,95);parseNavmeshJson(nav.json,nav.name);setSourceProgress(100,'重新加载完成：NavMesh');sourceStatusEl.textContent=`已读取磁盘最新 NavMesh：${nav.stamp}`;return}if(lastLevelpointsHandle){const level=await readJsonFromHandle('levelpoints','LevelPoints',0,95);parseLevelpointsJson(level.json,level.name);setSourceProgress(100,'重新加载完成：LevelPoints');sourceStatusEl.textContent=`已读取磁盘最新 LevelPoints：${level.stamp}`;return}if(lastNavmeshFile&&!lastNavmeshHandle)return promptReselectSnapshot('navmesh','NavMesh');if(lastLevelpointsFile&&!lastLevelpointsHandle)return promptReselectSnapshot('levelpoints','LevelPoints')}catch(e){sourceStatusEl.textContent=`加载失败：${e.message}`;setSourceProgress(0,'加载失败')}}
canvas.addEventListener('pointerdown',e=>{dragging=true;dragMoved=false;last={x:e.clientX,y:e.clientY};canvas.classList.add('dragging');canvas.setPointerCapture(e.pointerId)})
canvas.addEventListener('pointermove',e=>{const r=canvas.getBoundingClientRect(),sx=(e.clientX-r.left)*dpr,sy=(e.clientY-r.top)*dpr;if(dragging){if(Math.hypot(e.clientX-last.x,e.clientY-last.y)>1.5)dragMoved=true;const dx=(e.clientX-last.x)*dpr/canvas.width*view.w,dz=(e.clientY-last.y)*dpr/canvas.height*view.h;last={x:e.clientX,y:e.clientY};view.x-=dx;view.z-=dz;draw();return}if(!hoverEl.checked)return;if(!connectivityMode){const point=findPoint({x:sx,y:sy});if(point){tip.style.display='block';tip.style.left=`${Math.min(r.width-390,e.clientX-r.left+12)}px`;tip.style.top=`${Math.min(r.height-245,e.clientY-r.top+12)}px`;const issue=point.nav&&point.nav.issue?point.nav.issue:'OK';tip.innerHTML=`<b>${escapeHtml(rawTypeText(point))}</b><br>Issue: ${issue}<br>Label/Class: ${escapeHtml(point.label)} / ${escapeHtml(point.class_name)}<br>subID: ${point.sub_id??'-'} parent: ${point.parent_sub_id??'-'}<br>ObjID/Type: ${point.obj_id??'-'} / ${point.obj_type??'-'}<br>Display: ${escapeHtml(point.obj_display_name||'-')}<br>Npc/Team/Pet: ${point.npc_id??'-'} / ${point.team_id??'-'} / ${point.pet_monster_id??'-'}<br>AreaID: ${point.area_id??'-'} Layer: ${point.belong_to_layer_id??'-'}<br>X/Z/Y: ${fmt(point.x)}, ${fmt(point.z)}, ${fmt(point.y,3)}<br>Rotation: ${escapeHtml(rotationText(point))}<br>Forward: ${escapeHtml(forwardText(point))}<br>Poly: ${point.nav?point.nav.poly:'-'} Component: ${point.nav?point.nav.component:'-'}<br>${escapeHtml(point.prefab_path||'')}`;return}}const poly=findPoly(screenToWorld(sx,sy));if(!poly){tip.style.display='none';return}tip.style.display='block';tip.style.left=`${Math.min(r.width-280,e.clientX-r.left+12)}px`;tip.style.top=`${Math.min(r.height-155,e.clientY-r.top+12)}px`;tip.innerHTML=`<b>Polygon ${poly.i}</b><br>ProvinceID / INT_areaID: ${poly.province}<br>区域类型: ${escapeHtml(poly.areaKind||'未知区域')}<br>区域名: ${escapeHtml(poly.areaName||'-')}<br>gveType / pvpType: ${poly.gveType??'-'} / ${poly.pvpType??'-'}<br>AreaType(raw): ${poly.area}<br>Component: ${poly.component} (${poly.componentSize})<br>Vertexs: ${poly.verts.length}<br>Avg Y: ${fmt(poly.y,3)}`})
canvas.addEventListener('pointerup',e=>{if(!dragMoved){const r=canvas.getBoundingClientRect(),screen={x:(e.clientX-r.left)*dpr,y:(e.clientY-r.top)*dpr};if(connectivityMode){const poly=findPoly(screenToWorld(screen.x,screen.y));if(poly)runConnectivityCheck(poly);else statusEl.textContent='连通性检查：未命中 NavMesh，请点击任意 NavMesh 区域。'}else{const p=findPoint(screen);if(p)selectPoint(p)}}dragging=false;canvas.classList.remove('dragging');canvas.releasePointerCapture(e.pointerId)})
canvas.addEventListener('pointerleave',()=>{dragging=false;canvas.classList.remove('dragging');tip.style.display='none'})
canvas.addEventListener('wheel',e=>{e.preventDefault();const r=canvas.getBoundingClientRect();zoom(e.deltaY<0?.86:1.16,{x:(e.clientX-r.left)*dpr,y:(e.clientY-r.top)*dpr})},{passive:false})
function zoom(factor,screenPoint){const c=screenPoint||{x:canvas.width/2,y:canvas.height/2},before=screenToWorld(c.x,c.y);view.w*=factor;view.h*=factor;const after=screenToWorld(c.x,c.y);view.x+=before.x-after.x;view.z+=before.z-after.z;draw()}
document.getElementById('zoom-in').addEventListener('click',()=>zoom(.75));document.getElementById('zoom-out').addEventListener('click',()=>zoom(1.33));document.getElementById('reset').addEventListener('click',resetView);
async function chooseSourceFile(kind){if(window.levelpointsElectron){try{setSourceProgress(0,'Electron：选择根目录');const loaded=await window.levelpointsElectron.selectProjectRoot();if(!loaded)return;if(loaded.navmesh){parseNavmeshJson(loaded.navmesh.json,loaded.navmesh.name||loaded.navmesh.path)}if(loaded.levelpoints){parseLevelpointsJson(loaded.levelpoints.json,loaded.levelpoints.name||loaded.levelpoints.path)}setSourceProgress(100,'Electron：已按根目录加载默认相对路径文件');sourceStatusEl.textContent=`Electron 桌面版已设置根目录：${loaded.root||''}`;return}catch(e){sourceStatusEl.textContent=`选择根目录失败：${e.message}`;setSourceProgress(0,'选择根目录失败');return}}if(window.showOpenFilePicker){try{const [handle]=await window.showOpenFilePicker({types:[{description:'JSON files',accept:{'application/json':['.json']}}],multiple:false});await ensureHandlePermission(handle,kind==='navmesh'?'NavMesh':'LevelPoints');const file=await handle.getFile();const loaded=await readJsonFileAsync(file,kind==='navmesh'?'NavMesh':'LevelPoints',0,95);if(kind==='navmesh'){lastNavmeshHandle=handle;lastNavmeshFile=file;parseNavmeshJson(loaded.json,loaded.name);setSourceProgress(100,'已加载 NavMesh，并保存文件句柄用于刷新读取最新内容');sourceStatusEl.textContent=`已授权 NavMesh：${fileStamp(file)}。刷新会读取磁盘最新内容。`}else{lastLevelpointsHandle=handle;lastLevelpointsFile=file;parseLevelpointsJson(loaded.json,loaded.name);setSourceProgress(100,'已加载 LevelPoints，并保存文件句柄用于刷新读取最新内容');sourceStatusEl.textContent=`已授权 LevelPoints：${fileStamp(file)}。刷新会读取磁盘最新内容。`}return}catch(e){if(e&&e.name==='AbortError')return;sourceStatusEl.textContent=`选择文件失败：${e.message}`;setSourceProgress(0,'选择文件失败');return}}sourceStatusEl.textContent='当前浏览器不支持可持久刷新的文件句柄，只能重新选择文件来读取最新版本。';(kind==='navmesh'?navmeshFileEl:levelpointsFileEl).click()}
document.getElementById('choose-navmesh').addEventListener('click',()=>chooseSourceFile('navmesh'));document.getElementById('choose-levelpoints').addEventListener('click',()=>chooseSourceFile('levelpoints'));document.getElementById('refresh-source').addEventListener('click',refreshSource);
document.getElementById('run-validation').addEventListener('click',runValidationChecks);document.getElementById('clear-validation').addEventListener('click',clearValidationResults);
document.getElementById('run-station-distance').addEventListener('click',runStationDistanceCheck);document.getElementById('clear-station-distance').addEventListener('click',clearStationDistanceResults);
connectivityToggleEl.addEventListener('click',toggleConnectivityCheck);
rawTypeToggleEl.addEventListener('click',()=>{rawTypeTableEl.classList.toggle('collapsed');updateRawTypeToggle()});
filterSearchEl.addEventListener('input',()=>{filterQuery=filterSearchEl.value;renderFilters()});
filterAllEl.addEventListener('click',()=>{enabledGroups=new Set(filterEntries.map(e=>e.key));renderFilters();fillStats();draw()});
filterNoneEl.addEventListener('click',()=>{enabledGroups=new Set();renderFilters();fillStats();draw()});
provinceButtonEl.addEventListener('click',e=>{e.stopPropagation();provinceFilterEl.classList.toggle('open')});
provinceFilterEl.addEventListener('click',e=>e.stopPropagation());
document.addEventListener('click',()=>provinceFilterEl.classList.remove('open'));
provinceSearchEl.addEventListener('input',()=>{provinceQuery=provinceSearchEl.value;renderProvinceFilter()});
document.getElementById('province-all').addEventListener('click',()=>{selectedProvinces=new Set(allProvinceIds);renderProvinceFilter();fillStats();draw()});
document.getElementById('province-none').addEventListener('click',()=>{selectedProvinces=new Set();renderProvinceFilter();fillStats();draw()});
navmeshFileEl.addEventListener('change',async e=>{const f=e.target.files&&e.target.files[0];if(f){try{lastNavmeshFile=f;lastNavmeshHandle=null;const loaded=await readJsonFileAsync(f,'NavMesh',0,95);parseNavmeshJson(loaded.json,loaded.name);setSourceProgress(100,'已加载 NavMesh。当前浏览器未提供文件句柄，刷新会重读本次授权的 File 对象。')}catch(err){sourceStatusEl.textContent=`加载失败：${err.message}`;setSourceProgress(0,'加载失败')}}e.target.value=''});levelpointsFileEl.addEventListener('change',async e=>{const f=e.target.files&&e.target.files[0];if(f){try{lastLevelpointsFile=f;lastLevelpointsHandle=null;const loaded=await readJsonFileAsync(f,'LevelPoints',0,95);parseLevelpointsJson(loaded.json,loaded.name);setSourceProgress(100,'已加载 LevelPoints。当前浏览器未提供文件句柄，刷新会重读本次授权的 File 对象。')}catch(err){sourceStatusEl.textContent=`加载失败：${err.message}`;setSourceProgress(0,'加载失败')}}e.target.value=''});
baseModeEl.addEventListener('change',draw);modeEl.addEventListener('change',draw);rotationEl.addEventListener('change',resetView);mirrorEl.addEventListener('change',resetView);searchEl.addEventListener('input',()=>{fillStats();draw()});strokeEl.addEventListener('change',draw);pointsEl.addEventListener('change',()=>{fillStats();draw()});issuesEl.addEventListener('change',()=>{fillStats();draw()});window.addEventListener('resize',resize);
setupElectronControls();rebuildIndexes();rebuildProvinceFilter(true);rebuildFilters(true);fillStats();resetView();resize();window.__viewerDiagnostics=()=>({title:document.title,bodyText:document.body.innerText.slice(0,500),polygonCount:polygons.length,boundarySegmentCount:boundarySegments.length,pointCount:points.length,provinceCount:allProvinceIds.length,selectedProvinceCount:selectedProvinces.size,statsSource:stats.source,levelpointsSource:stats.levelpointsSource,errors:window.__viewerErrors||[]});
</script>
</body>
</html>"""
    Path(args.output).parent.mkdir(parents=True, exist_ok=True)
    Path(args.output).write_text(html.replace("__PAYLOAD__", payload).replace("__TOP_BANNER__", top_banner_uri), encoding="utf-8")


if __name__ == "__main__":
    main()
