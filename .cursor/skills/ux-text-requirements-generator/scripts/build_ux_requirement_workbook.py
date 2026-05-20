from __future__ import annotations

import json
import math
import os
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from PIL import Image


def long_path(path: str | Path) -> str:
    normalized = os.path.normpath(str(path))
    return normalized if normalized.startswith("\\\\?\\") else "\\\\?\\" + normalized


def load_config(config_path: Path) -> dict:
    return json.loads(config_path.read_text(encoding="utf-8"))


def crop_images(config: dict, root: Path) -> dict[str, Path]:
    crop_dir = root / "tmp" / config.get("crop_dir_name", "ux_requirement_crops")
    crop_dir.mkdir(parents=True, exist_ok=True)

    output_paths: dict[str, Path] = {}
    for section in config["sections"]:
        source_image_path = section.get("source_image", config["source_image"])
        image = Image.open(long_path(source_image_path))
        actual_width, actual_height = image.size
        coord_width, coord_height = section.get(
            "source_coordinate_size",
            config.get("source_coordinate_size", [actual_width, actual_height]),
        )
        scale_x = actual_width / coord_width if coord_width else 1
        scale_y = actual_height / coord_height if coord_height else 1
        destination = crop_dir / section["crop_name"]
        x1, y1, x2, y2 = section["crop_box"]
        scaled_box = (
            round(x1 * scale_x),
            round(y1 * scale_y),
            round(x2 * scale_x),
            round(y2 * scale_y),
        )
        crop = image.crop(scaled_box)
        crop.save(destination)
        output_paths[section["title"]] = destination
    return output_paths


def resolve_image_display_size(config: dict, section: dict, img: XLImage) -> tuple[int, int]:
    auto_image_width = section.get("auto_image_width", config.get("auto_image_width", False))
    if auto_image_width:
        max_image_width = int(section.get("max_image_width", config.get("max_image_width", 560)))
        min_image_width = int(section.get("min_image_width", config.get("min_image_width", 360)))
        image_width = min(max_image_width, img.width) if img.width else max_image_width
        if img.width and img.width < min_image_width:
            image_width = img.width
    else:
        image_width = int(section.get("image_width", config.get("image_width", 260)))

    scale = image_width / img.width if img.width else 1
    image_height = int(math.ceil(img.height * scale))
    return image_width, image_height


def resolve_reserved_rows(config: dict, section: dict, image_height: int) -> int:
    configured_rows = int(section.get("reserved_rows", 0))
    pixels_per_row = int(section.get("pixels_per_row", config.get("pixels_per_row", 20)))
    image_padding_rows = int(section.get("image_padding_rows", config.get("image_padding_rows", 2)))
    auto_rows = int(math.ceil(image_height / max(pixels_per_row, 1))) + image_padding_rows
    return max(configured_rows, auto_rows)


def apply_layout(ws) -> None:
    widths = {
        "A": 18,
        "B": 18,
        "C": 18,
        "D": 18,
        "E": 18,
        "F": 4,
        "G": 16,
        "H": 16,
        "I": 16,
        "J": 4,
        "K": 22,
        "L": 22,
        "M": 20,
        "N": 20,
        "O": 20,
        "P": 20,
        "Q": 20,
    }
    for column, width in widths.items():
        ws.column_dimensions[column].width = width
    ws.sheet_view.showGridLines = True


def style_sheet(ws) -> None:
    thin_gray = Side(style="thin", color="999999")
    border = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)
    title_fill = PatternFill("solid", fgColor="D9EAF7")
    header_fill = PatternFill("solid", fgColor="7030A0")
    header_font = Font(color="FFFFFF", bold=True)
    title_font = Font(bold=True)

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if cell.column >= 11:
                cell.border = border

    for row in ws.iter_rows():
        if ws.cell(row=row[0].row, column=11).value == "key":
            for col in range(1, 10):
                ws.cell(row=row[0].row, column=col).fill = title_fill
                ws.cell(row=row[0].row, column=col).font = title_font
            for col in range(11, 18):
                ws.cell(row=row[0].row, column=col).fill = header_fill
                ws.cell(row=row[0].row, column=col).font = header_font


def add_overview_section(ws, config: dict, start_row: int) -> int:
    overview = config.get("overview_rows", [])
    if not overview:
        return start_row

    title_row = start_row
    ws.merge_cells(start_row=title_row, start_column=1, end_row=title_row, end_column=17)
    ws.cell(title_row, 1).value = config.get("overview_title", "UI结构总览")
    for col in range(1, 18):
        ws.cell(title_row, col).fill = PatternFill("solid", fgColor="D9EAF7")
        ws.cell(title_row, col).font = Font(bold=True)

    note_row = title_row + 1
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=17)
    ws.cell(note_row, 1).value = config.get(
        "overview_note",
        "先阅读整体界面结构和流程分段，再进入逐页截图与文本需求。",
    )

    header_row = note_row + 1
    ws.merge_cells(start_row=header_row, start_column=1, end_row=header_row, end_column=4)
    ws.merge_cells(start_row=header_row, start_column=5, end_row=header_row, end_column=10)
    ws.merge_cells(start_row=header_row, start_column=11, end_row=header_row, end_column=17)
    ws.cell(header_row, 1).value = "模块"
    ws.cell(header_row, 5).value = "界面清单"
    ws.cell(header_row, 11).value = "阅读说明"
    for col in range(1, 18):
        ws.cell(header_row, col).fill = PatternFill("solid", fgColor="7030A0")
        ws.cell(header_row, col).font = Font(color="FFFFFF", bold=True)

    current_row = header_row + 1
    for item in overview:
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)
        ws.merge_cells(start_row=current_row, start_column=5, end_row=current_row, end_column=10)
        ws.merge_cells(start_row=current_row, start_column=11, end_row=current_row, end_column=17)
        ws.cell(current_row, 1).value = item.get("module", "")
        ws.cell(current_row, 5).value = item.get("pages", "")
        ws.cell(current_row, 11).value = item.get("notes", "")
        ws.row_dimensions[current_row].height = 34
        current_row += 1

    return current_row + 1


def build_workbook(config: dict, crop_paths: dict[str, Path]) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = config.get("sheet_title", "文本键值")
    apply_layout(ws)

    current_row = add_overview_section(ws, config, 1)
    for section in config["sections"]:
        title_row = current_row
        ws.merge_cells(start_row=title_row, start_column=1, end_row=title_row, end_column=9)
        ws.merge_cells(start_row=title_row, start_column=11, end_row=title_row, end_column=12)
        ws.merge_cells(start_row=title_row, start_column=13, end_row=title_row, end_column=14)
        ws.merge_cells(start_row=title_row, start_column=15, end_row=title_row, end_column=17)
        ws.cell(title_row, 1).value = section["title"]
        ws.cell(title_row, 11).value = "key"
        ws.cell(title_row, 13).value = "内容"
        ws.cell(title_row, 15).value = "备注"

        note_row = title_row + 1
        ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=9)
        ws.cell(note_row, 1).value = section["image_note"]

        image_cell = section.get("image_anchor", f"A{note_row + 1}")
        img = XLImage(str(crop_paths[section["title"]]))
        image_width, image_height = resolve_image_display_size(config, section, img)
        img.width = image_width
        img.height = image_height
        ws.add_image(img, image_cell)

        reserved_rows = resolve_reserved_rows(config, section, image_height)
        for row_idx in range(note_row + 1, note_row + 1 + reserved_rows):
            ws.row_dimensions[row_idx].height = 22

        right_start_row = note_row + 2
        for offset, item in enumerate(section["rows"]):
            row_idx = right_start_row + offset
            ws.merge_cells(start_row=row_idx, start_column=7, end_row=row_idx, end_column=9)
            ws.cell(row_idx, 7).value = item["note"]
            if item.get("key"):
                ws.merge_cells(start_row=row_idx, start_column=11, end_row=row_idx, end_column=12)
                ws.merge_cells(start_row=row_idx, start_column=13, end_row=row_idx, end_column=14)
                ws.merge_cells(start_row=row_idx, start_column=15, end_row=row_idx, end_column=17)
                ws.cell(row_idx, 11).value = item["key"]
                ws.cell(row_idx, 13).value = item.get("content", "")
                ws.cell(row_idx, 15).value = item.get("remark", item["note"])

        current_row = max(note_row + reserved_rows + 1, right_start_row + len(section["rows"])) + 1

    style_sheet(ws)
    return wb


def main() -> None:
    if len(sys.argv) != 2:
        raise SystemExit("Usage: py -3 tools/build_ux_requirement_workbook.py <config.json>")

    root = Path(__file__).resolve().parents[1]
    config_path = Path(sys.argv[1])
    if not config_path.is_absolute():
        config_path = root / config_path

    config = load_config(config_path)
    crop_paths = crop_images(config, root)
    workbook = build_workbook(config, crop_paths)
    output_path = Path(config["output_path"])
    if not output_path.is_absolute():
        output_path = root / output_path
    workbook.save(output_path)
    print(output_path)


if __name__ == "__main__":
    main()
