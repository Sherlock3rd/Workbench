from __future__ import annotations

import json
import math
import tempfile
from pathlib import Path

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from PIL import Image


THIN = Side(style="thin", color="C9D2E3")
BOX_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
TITLE_FILL = PatternFill("solid", fgColor="1F4E78")
SUBTITLE_FILL = PatternFill("solid", fgColor="D9EAF7")
TABLE_HEADER_FILL = PatternFill("solid", fgColor="EAF2F8")


def read_config(config_path: Path) -> dict:
    return json.loads(config_path.read_text(encoding="utf-8"))


def ensure_parent(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)


def autosize_columns(ws, widths: dict[str, int]) -> None:
    for column, width in widths.items():
        ws.column_dimensions[column].width = width


def row_span_from_image(image_path: Path, target_width_px: int, pixels_per_row: int) -> int:
    with Image.open(image_path) as img:
        width, height = img.size
    scale = min(1.0, target_width_px / max(1, width))
    rendered_height = math.ceil(height * scale)
    return max(8, math.ceil(rendered_height / max(1, pixels_per_row)))


def build_crop(source_image: Path, crop_box: list[int], crop_dir: Path, crop_name: str) -> Path:
    crop_dir.mkdir(parents=True, exist_ok=True)
    crop_path = crop_dir / crop_name
    with Image.open(source_image) as img:
        cropped = img.crop(tuple(crop_box))
        cropped.save(crop_path)
    return crop_path


def resolve_section_image(config: dict, section: dict, crop_dir: Path) -> Path | None:
    image_path = section.get("image_path")
    if image_path:
        return Path(image_path)

    source_image = config.get("source_image")
    crop_box = section.get("crop_box")
    crop_name = section.get("crop_name")
    if source_image and crop_box and crop_name:
        return build_crop(Path(source_image), crop_box, crop_dir, crop_name)

    return None


def apply_base_style(ws) -> None:
    autosize_columns(
        ws,
        {
            "A": 16,
            "B": 16,
            "C": 16,
            "D": 16,
            "E": 16,
            "F": 3,
            "G": 20,
            "H": 20,
            "I": 20,
            "J": 3,
            "K": 28,
            "L": 32,
            "M": 42,
        },
    )
    ws.sheet_view.showGridLines = False


def write_merged_text(ws, cell_range: str, value: str, *, font: Font | None = None, fill=None, alignment=None):
    ws.merge_cells(cell_range)
    cell = ws[cell_range.split(":")[0]]
    cell.value = value
    cell.border = BOX_BORDER
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if alignment:
        cell.alignment = alignment


def add_image(ws, image_path: Path, anchor: str, target_width_px: int) -> None:
    image = XLImage(str(image_path))
    if image.width > target_width_px:
        scale = target_width_px / image.width
        image.width = int(image.width * scale)
        image.height = int(image.height * scale)
    ws.add_image(image, anchor)


def write_overview(ws, config: dict, start_row: int) -> int:
    current_row = start_row
    write_merged_text(
        ws,
        f"A{current_row}:M{current_row}",
        config.get("overview_title", "UI结构总览"),
        font=Font(bold=True, color="FFFFFF", size=14),
        fill=TITLE_FILL,
        alignment=Alignment(horizontal="left", vertical="center"),
    )
    ws.row_dimensions[current_row].height = 24
    current_row += 1

    write_merged_text(
        ws,
        f"A{current_row}:M{current_row}",
        config.get("overview_note", ""),
        alignment=Alignment(wrap_text=True, vertical="top"),
    )
    ws.row_dimensions[current_row].height = 42
    current_row += 1

    headers = [("A", "模块"), ("D", "页面"), ("H", "阅读说明")]
    for col, title in headers:
        cell = ws[f"{col}{current_row}"]
        cell.value = title
        cell.font = Font(bold=True)
        cell.fill = TABLE_HEADER_FILL
        cell.border = BOX_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(f"A{current_row}:C{current_row}")
    ws.merge_cells(f"D{current_row}:G{current_row}")
    ws.merge_cells(f"H{current_row}:M{current_row}")
    current_row += 1

    for row in config.get("overview_rows", []):
        ws.merge_cells(f"A{current_row}:C{current_row}")
        ws.merge_cells(f"D{current_row}:G{current_row}")
        ws.merge_cells(f"H{current_row}:M{current_row}")
        values = {
            f"A{current_row}": row.get("module", ""),
            f"D{current_row}": row.get("pages", ""),
            f"H{current_row}": row.get("notes", ""),
        }
        for address, value in values.items():
            cell = ws[address]
            cell.value = value
            cell.border = BOX_BORDER
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[current_row].height = 48
        current_row += 1

    return current_row + 1


def write_section(ws, config: dict, section: dict, start_row: int, crop_dir: Path) -> int:
    current_row = start_row
    write_merged_text(
        ws,
        f"A{current_row}:M{current_row}",
        section.get("title", "未命名章节"),
        font=Font(bold=True, color="FFFFFF", size=13),
        fill=TITLE_FILL,
        alignment=Alignment(horizontal="left", vertical="center"),
    )
    ws.row_dimensions[current_row].height = 24
    current_row += 1

    image_target_width = int(config.get("max_image_width", 560))
    pixels_per_row = int(config.get("pixels_per_row", 20))
    section_image = resolve_section_image(config, section, crop_dir)
    image_rows = 14
    if section_image and section_image.exists():
        image_rows = row_span_from_image(section_image, image_target_width, pixels_per_row)
        add_image(ws, section_image, f"A{current_row}", image_target_width)

    note_row_end = current_row + max(image_rows - 1, 5)
    ws.merge_cells(f"G{current_row}:I{note_row_end}")
    note_cell = ws[f"G{current_row}"]
    note_cell.value = section.get("image_note", "")
    note_cell.border = BOX_BORDER
    note_cell.alignment = Alignment(wrap_text=True, vertical="top")

    header_row = current_row
    for column, value in [("K", "交互说明"), ("L", "key"), ("M", "内容 / 备注")]:
        cell = ws[f"{column}{header_row}"]
        cell.value = value
        cell.font = Font(bold=True)
        cell.fill = SUBTITLE_FILL
        cell.border = BOX_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")

    rows = section.get("rows", [])
    body_row = current_row + 1
    for row in rows:
        values = {
            f"K{body_row}": row.get("note", ""),
            f"L{body_row}": row.get("key", ""),
            f"M{body_row}": f"{row.get('content', '')}\n备注：{row.get('remark', '')}".strip(),
        }
        for address, value in values.items():
            cell = ws[address]
            cell.value = value
            cell.border = BOX_BORDER
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[body_row].height = 54
        body_row += 1

    current_row = max(note_row_end, body_row - 1) + int(config.get("image_padding_rows", 2)) + 1
    return current_row


def build_workbook(config_path: Path) -> Path:
    config = read_config(config_path)
    output_path = (config_path.parent / config.get("output_path", "ux_text_requirements.xlsx")).resolve()
    ensure_parent(output_path)

    crop_dir_name = config.get("crop_dir_name", "ux_requirement_crops")
    crop_dir = output_path.parent / crop_dir_name

    workbook = Workbook()
    sheet_title = config.get("sheet_title", "UxTextRequirements")
    ws = workbook.active
    ws.title = sheet_title[:31]
    apply_base_style(ws)

    row = 1
    row = write_overview(ws, config, row)
    for section in config.get("sections", []):
        row = write_section(ws, config, section, row, crop_dir)

    workbook.save(output_path)
    return output_path


def main() -> int:
    import argparse

    parser = argparse.ArgumentParser(description="Build a screenshot-backed UX requirement workbook from a JSON config.")
    parser.add_argument("config", help="Path to the workbook config JSON")
    args = parser.parse_args()

    output_path = build_workbook(Path(args.config))
    print(output_path)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
